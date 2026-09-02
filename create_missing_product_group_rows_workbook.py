from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
OUTPUTS = BASE / "outputs"
SOURCE_CSV = OUTPUTS / "Innoflame_merged_sales_csv_source_with_L1_L2_L3_delivery_handling.csv"
OUTPUT_XLSX = OUTPUTS / "missing_product_group_rows.xlsx"


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 250)):
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                current_width = ws.column_dimensions[get_column_letter(cell.column)].width or 0
                ws.column_dimensions[get_column_letter(cell.column)].width = min(
                    max(current_width, len(value) + 2),
                    45,
                )

    wb.save(path)
    wb.close()


def main() -> None:
    df = pd.read_csv(SOURCE_CSV, dtype=str, encoding="utf-8-sig", keep_default_na=False)
    missing_mask = df["product_group_l3_code"].fillna("").str.strip().eq("")
    missing = df.loc[missing_mask].copy()

    sales_numeric = pd.to_numeric(
        missing["sales"].astype(str).str.replace(",", ".", regex=False),
        errors="coerce",
    ).fillna(0)

    summary = pd.DataFrame(
        [
            {"metric": "source_rows_total", "value": int(len(df))},
            {"metric": "missing_product_group_rows", "value": int(len(missing))},
            {"metric": "missing_product_group_unique_names", "value": int(missing["name"].nunique())},
            {"metric": "missing_product_group_sales_eur", "value": round(float(sales_numeric.sum()), 2)},
        ]
    )

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        missing.to_excel(writer, index=False, sheet_name="missing_rows")
        summary.to_excel(writer, index=False, sheet_name="summary")

    style_workbook(OUTPUT_XLSX)

    wb = load_workbook(OUTPUT_XLSX, read_only=True, data_only=True)
    sheet_info = {ws.title: {"rows": ws.max_row, "cols": ws.max_column} for ws in wb.worksheets}
    wb.close()

    result = {
        "source_csv": str(SOURCE_CSV.resolve()),
        "output_xlsx": str(OUTPUT_XLSX.resolve()),
        "missing_rows": int(len(missing)),
        "sheet_info": sheet_info,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
