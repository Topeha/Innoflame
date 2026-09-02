from __future__ import annotations

import argparse
import importlib.util
import json
import logging
import re
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_CRM_PATH = BASE_DIR / "CRM_potentials_03.06.2026_03.07.2026 (1).xlsx"
DEFAULT_GROUPING_PATH = BASE_DIR / "product_master_enrichment" / "final_product_grouping" / "Innoflame_tuoteryhmittely.xlsx"
DEFAULT_OUTPUT_XLSX = BASE_DIR / "current_customer_potential_with_product_groups.xlsx"
DEFAULT_CURRENT_CUSTOMER_CSV = BASE_DIR / "current_customer_potential.csv"
DEFAULT_RECOMMENDATIONS_CSV = BASE_DIR / "product_group_recommendations.csv"
DEFAULT_VALIDATION_CSV = BASE_DIR / "validation_against_crm.csv"
DEFAULT_ORIGINAL_MODEL = BASE_DIR / "prospect_model.py"
DEFAULT_V3_MODEL = BASE_DIR / "two_stage_potential_model" / "v3_recent_weighted_current_model" / "innoflame_all_accounts_model_v3.py"
CRM_COLUMNS_F_TO_K = ["Type", "Status", "Sales", "Probability", "ETA", "Comment"]

LOGGER = logging.getLogger(__name__)


def normalize_business_id(value: Any) -> str | None:
    if pd.isna(value):
        return None
    if isinstance(value, (int, np.integer)):
        text = str(int(value))
    elif isinstance(value, (float, np.floating)) and float(value).is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if not text:
        return None
    if text.upper().startswith("FI"):
        text = text[2:]
    digits = "".join(character for character in text if character.isdigit())
    if len(digits) == 7:
        digits = f"0{digits}"
    if len(digits) >= 8:
        return f"{digits[:-1]}-{digits[-1]}"
    if re.fullmatch(r"\d{7,8}-\d", text):
        return text
    return None


def normalize_column_name(value: Any) -> str:
    text = str(value).strip().lower()
    replacements = str.maketrans({"ä": "a", "ö": "o", "å": "a"})
    text = text.translate(replacements)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def normalize_name(value: Any) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip().lower()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.translate(str.maketrans({"ä": "a", "ö": "o", "å": "a"}))
    text = re.sub(r"\b(oyj|oy|ab|ltd|limited|inc|gmbh|as)\b", " ", text)
    text = re.sub(r"\b(suljettu|closed|myynti|gokeep|hr lahjat|ei saa laskuttaa|ylitoimituksia)\b", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = " ".join(text.split())
    return text or None


def resolve_existing_column(frame: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized = {normalize_column_name(column): column for column in frame.columns}
    for candidate in candidates:
        value = normalized.get(normalize_column_name(candidate))
        if value is not None:
            return value
    return None


def read_inputs(args: argparse.Namespace) -> dict[str, pd.DataFrame]:
    crm = pd.read_excel(args.crm_potentials, sheet_name=0)
    grouping = pd.read_excel(args.product_grouping, sheet_name="Products")
    accounts = normalize_accounts_source(pd.read_excel(args.accounts))
    sales = pd.read_csv(args.sales)
    companies = pd.read_excel(args.companies)
    return {
        "crm": crm,
        "product_grouping": grouping,
        "accounts": accounts,
        "sales": sales,
        "companies": companies,
    }


def normalize_accounts_source(accounts: pd.DataFrame) -> pd.DataFrame:
    frame = accounts.copy()
    rename_map = {
        "id": "ID",
        "name": "Account Name",
        "company_name": "Company Name",
        "business_id": "Business ID",
        "category": "customer_status",
    }
    for source, target in rename_map.items():
        if source in frame.columns and target not in frame.columns:
            frame[target] = frame[source]
    if "Business ID" not in frame.columns:
        business_id_col = resolve_existing_column(frame, ["business id", "business_id", "y tunnus", "y-tunnus"])
        if business_id_col:
            frame["Business ID"] = frame[business_id_col]
    if "ID" not in frame.columns:
        account_id_col = resolve_existing_column(frame, ["id", "account_id", "account id"])
        if account_id_col:
            frame["ID"] = frame[account_id_col]
    if "Company Name" not in frame.columns:
        company_col = resolve_existing_column(frame, ["company name", "company_name", "name", "account name"])
        if company_col:
            frame["Company Name"] = frame[company_col]
    if "Account Name" not in frame.columns:
        frame["Account Name"] = frame.get("Company Name", frame.get("Business ID"))
    if "customer_status" not in frame.columns:
        status_col = resolve_existing_column(frame, ["customer status", "customer_status", "category"])
        frame["customer_status"] = frame[status_col] if status_col else "Active"
    return frame


def detect_product_group_columns(grouping: pd.DataFrame) -> dict[int, dict[str, str | None]]:
    result: dict[int, dict[str, str | None]] = {}
    normalized = {normalize_column_name(column): column for column in grouping.columns}
    for level in range(1, 8):
        code_candidates = [
            f"product group l{level} code",
            f"product_group_l{level}_code",
            f"l{level} code",
            f"tuoteryhma l{level} koodi",
            f"tuoteryhma taso {level} koodi",
        ]
        name_candidates = [
            f"product group l{level} name",
            f"product_group_l{level}_name",
            f"l{level} name",
            f"tuoteryhma l{level} nimi",
            f"tuoteryhma taso {level} nimi",
        ]
        code_col = next((normalized.get(normalize_column_name(candidate)) for candidate in code_candidates if normalized.get(normalize_column_name(candidate))), None)
        name_col = next((normalized.get(normalize_column_name(candidate)) for candidate in name_candidates if normalized.get(normalize_column_name(candidate))), None)
        if code_col or name_col:
            result[level] = {"code": code_col, "name": name_col}
    if not result:
        raise ValueError("No product group level columns were found in the product grouping file.")
    return result


def create_lowest_product_group(grouping: pd.DataFrame) -> tuple[pd.DataFrame, dict[int, dict[str, str | None]]]:
    frame = grouping.copy()
    level_columns = detect_product_group_columns(frame)
    frame["lowest_product_group_level"] = pd.NA
    frame["lowest_product_group_code"] = pd.NA
    frame["lowest_product_group_name"] = pd.NA

    for level in sorted(level_columns, reverse=True):
        code_col = level_columns[level]["code"]
        name_col = level_columns[level]["name"]
        code = frame[code_col].astype("string").str.strip() if code_col else pd.Series(pd.NA, index=frame.index, dtype="string")
        name = frame[name_col].astype("string").str.strip() if name_col else pd.Series(pd.NA, index=frame.index, dtype="string")
        has_value = code.fillna("").ne("") | name.fillna("").ne("")
        target = frame["lowest_product_group_code"].isna() & has_value
        frame.loc[target, "lowest_product_group_level"] = f"L{level}"
        frame.loc[target, "lowest_product_group_code"] = code[target].where(code[target].fillna("").ne(""), name[target])
        frame.loc[target, "lowest_product_group_name"] = name[target].where(name[target].fillna("").ne(""), code[target])

    return frame, level_columns


def load_module(path: Path, module_name: str) -> Any:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load module from {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_model_artifacts(args: argparse.Namespace, inputs: dict[str, pd.DataFrame]) -> dict[str, Any]:
    original_model = load_module(Path(args.original_model), "innoflame_original_prospect_model_current_customer")
    v3_model = load_module(Path(args.v3_model), "innoflame_v3_current_customer_adapter")
    original_model.configure_logging()

    def weighted_targets_adapter(
        sales: pd.DataFrame,
        accounts: pd.DataFrame,
        lookback_days: int,
        min_training_customer_annual_sales_eur: float,
    ) -> tuple[pd.DataFrame, pd.Timestamp]:
        return v3_model.compute_weighted_customer_targets(
            sales=sales,
            accounts=accounts,
            lookback_days=lookback_days,
            min_training_customer_annual_sales_eur=min_training_customer_annual_sales_eur,
            recent_year_weight=args.recent_year_weight,
            middle_year_weight=args.middle_year_weight,
            oldest_year_weight=args.oldest_year_weight,
        )

    original_model.compute_customer_targets = weighted_targets_adapter
    scoring_universe = v3_model.build_scoring_universe(inputs["companies"], inputs["accounts"], original_model)
    modeling_df, top_customers, reference_date = original_model.build_modeling_frame(
        accounts=inputs["accounts"],
        sales=inputs["sales"],
        companies=scoring_universe,
        top_n_customers=args.top_n_customers,
        lookback_days=args.lookback_days,
        min_training_customer_annual_sales_eur=args.min_training_customer_annual_sales_eur,
    )
    numeric_features, categorical_features = original_model.feature_columns()
    trained_model, metrics = original_model.train_model(
        modeling_df=modeling_df,
        numeric_features=numeric_features,
        categorical_features=categorical_features,
        random_state=args.random_state,
    )
    account_frame = original_model.preprocess_accounts(inputs["accounts"])
    excluded_business_ids = original_model.load_excluded_business_ids(args.exclude_business_ids_file)
    all_scored = v3_model.score_all_accounts(
        trained_model=trained_model,
        modeling_df=modeling_df,
        account_frame=account_frame,
        top_customers=top_customers,
        excluded_business_ids=excluded_business_ids,
        reference_date=reference_date,
        model=original_model,
        current_customer_recent_sales_weight=args.current_customer_recent_sales_weight,
        recent_sales_floor_multiplier=args.recent_sales_floor_multiplier,
    )
    all_scored = all_scored.rename(columns={"ennustettu potentiaali": "estimated_potential_eur"})
    if "final_value_eur" not in all_scored.columns:
        all_scored["final_value_eur"] = all_scored["estimated_potential_eur"]
    return {
        "original_model": original_model,
        "v3_model": v3_model,
        "modeling_df": modeling_df,
        "top_customers": top_customers,
        "reference_date": reference_date,
        "trained_model": trained_model,
        "metrics": metrics,
        "account_frame": account_frame,
        "all_scored": all_scored,
        "feature_columns": numeric_features + categorical_features,
        "recent_year_weight": args.recent_year_weight,
        "middle_year_weight": args.middle_year_weight,
        "oldest_year_weight": args.oldest_year_weight,
        "current_customer_recent_sales_weight": args.current_customer_recent_sales_weight,
        "recent_sales_floor_multiplier": args.recent_sales_floor_multiplier,
    }


def prepare_customer_features(crm: pd.DataFrame, accounts: pd.DataFrame, scored: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    frame = crm.copy()
    frame["_input_row_id"] = np.arange(len(frame), dtype=int) + 1
    key_col = resolve_existing_column(frame, ["business_id", "business id", "y tunnus", "y-tunnus", "ytunnus"])
    account_id_col = resolve_existing_column(frame, ["account_id", "account id", "customer_id", "customer id", "id"])
    name_col = resolve_existing_column(frame, ["name", "account name", "company name", "customer", "asiakas"])

    if key_col:
        frame["business_id"] = frame[key_col].map(normalize_business_id)
    else:
        frame["business_id"] = pd.NA
    if account_id_col:
        frame["account_id"] = pd.to_numeric(frame[account_id_col], errors="coerce")
    else:
        frame["account_id"] = pd.NA
    if name_col:
        frame["_normalized_name"] = frame[name_col].map(normalize_name)
    else:
        frame["_normalized_name"] = pd.NA

    account_lookup = accounts.copy()
    account_business_col = resolve_existing_column(account_lookup, ["business id", "business_id", "y tunnus", "y-tunnus"])
    account_id_source = resolve_existing_column(account_lookup, ["id", "account_id", "account id"])
    account_name_col = resolve_existing_column(account_lookup, ["account name", "company name", "virallinen nimi", "name"])
    account_lookup["_account_business_id"] = account_lookup[account_business_col].map(normalize_business_id) if account_business_col else pd.NA
    account_lookup["_account_id"] = pd.to_numeric(account_lookup[account_id_source], errors="coerce") if account_id_source else pd.NA
    account_lookup["_normalized_name"] = account_lookup[account_name_col].map(normalize_name) if account_name_col else pd.NA

    by_account_id = (
        account_lookup.dropna(subset=["_account_id", "_account_business_id"])
        .drop_duplicates(subset=["_account_id"])
        .set_index("_account_id")["_account_business_id"]
        .to_dict()
    )
    missing_business = frame["business_id"].isna() & frame["account_id"].notna()
    frame.loc[missing_business, "business_id"] = frame.loc[missing_business, "account_id"].map(by_account_id)

    by_name = (
        account_lookup.dropna(subset=["_normalized_name", "_account_business_id"])
        .drop_duplicates(subset=["_normalized_name"])
        .set_index("_normalized_name")["_account_business_id"]
        .to_dict()
    )
    missing_business = frame["business_id"].isna() & frame["_normalized_name"].notna()
    frame.loc[missing_business, "business_id"] = frame.loc[missing_business, "_normalized_name"].map(by_name)

    account_name_pairs = (
        account_lookup.dropna(subset=["_normalized_name", "_account_business_id"])
        .drop_duplicates(subset=["_normalized_name", "_account_business_id"])
        .loc[:, ["_normalized_name", "_account_business_id"]]
        .itertuples(index=False, name=None)
    )
    account_name_pairs = [(str(name), business_id) for name, business_id in account_name_pairs if len(str(name)) >= 4]
    for row_index, crm_name in frame.loc[frame["business_id"].isna(), "_normalized_name"].dropna().items():
        crm_name = str(crm_name)
        if len(crm_name) < 4:
            continue
        matches = {
            business_id
            for account_name, business_id in account_name_pairs
            if crm_name in account_name or account_name in crm_name
        }
        if len(matches) == 1:
            frame.at[row_index, "business_id"] = next(iter(matches))

    scored_keys = scored[["business_id"]].dropna().drop_duplicates().copy()
    matched = frame.merge(scored_keys.assign(_has_model_score=True), on="business_id", how="left")
    matched["_has_model_score"] = matched["_has_model_score"].eq(True)
    return frame, matched


def score_current_customers(crm_features: pd.DataFrame, scored: pd.DataFrame) -> pd.DataFrame:
    score_columns = [
        "business_id",
        "rank",
        "priority",
        "company",
        "score",
        "segment_median_value_eur",
        "model_value_eur",
        "baseline_value_eur",
        "pre_sales_adjusted_final_value_eur",
        "recent_sales_value_eur",
        "sales_history_adjustment_eur",
        "final_value_eur",
        "probability_of_growth",
        "sales_momentum_ratio",
        "conditional_potential_eur",
        "expected_potential_eur",
        "estimated_potential_eur",
        "avg_annual_sales_3y_eur",
        "avg_annual_sales_3y_eur_unweighted",
        "recent_12m",
        "middle_12m",
        "oldest_12m",
        "revenue_k_eur",
        "company_segment",
        "segment_lift",
        "industry",
        "positive_signals",
        "is_account_customer",
        "reference_date",
    ]
    available = [column for column in score_columns if column in scored.columns]
    merged = crm_features.merge(scored[available].drop_duplicates(subset=["business_id"]), on="business_id", how="left")
    merged["model_estimated_potential_eur"] = pd.to_numeric(merged.get("estimated_potential_eur"), errors="coerce")
    merged["customer_potential_rank"] = merged["model_estimated_potential_eur"].rank(method="first", ascending=False, na_option="bottom")
    merged["customer_potential_rank"] = merged["customer_potential_rank"].astype("Int64")
    return merged


def collapse_to_one_row_per_customer(customer_potential: pd.DataFrame) -> pd.DataFrame:
    frame = customer_potential.copy()
    if "business_id" not in frame.columns:
        frame["business_id"] = pd.NA
    if "_normalized_name" not in frame.columns:
        frame["_normalized_name"] = frame.get("Name", pd.Series(pd.NA, index=frame.index)).map(normalize_name)

    business_key = frame["business_id"].astype("string")
    name_key = frame["_normalized_name"].astype("string")
    row_key = frame.get("_input_row_id", pd.Series(np.arange(len(frame)) + 1, index=frame.index)).astype("string")
    frame["_customer_key"] = np.where(
        business_key.fillna("").ne(""),
        "bid:" + business_key.fillna(""),
        np.where(name_key.fillna("").ne(""), "name:" + name_key.fillna(""), "row:" + row_key.fillna("")),
    )

    sales_col = resolve_existing_column(frame, ["sales"])
    probability_col = resolve_existing_column(frame, ["probability"])
    crm_sales = pd.to_numeric(frame[sales_col], errors="coerce") if sales_col else pd.Series(0.0, index=frame.index)
    probability = pd.to_numeric(frame[probability_col], errors="coerce") if probability_col else pd.Series(1.0, index=frame.index)
    probability = probability.where(probability.le(1.0), probability / 100.0).fillna(1.0)
    frame["_crm_row_potential_eur"] = crm_sales.fillna(0.0) * probability

    frame["_has_model"] = pd.to_numeric(frame.get("model_estimated_potential_eur"), errors="coerce").notna()
    frame["_sort_model_potential"] = pd.to_numeric(frame.get("model_estimated_potential_eur"), errors="coerce").fillna(-1.0)
    frame["_sort_input_row"] = pd.to_numeric(frame.get("_input_row_id"), errors="coerce").fillna(np.inf)
    ordered = frame.sort_values(
        ["_customer_key", "_has_model", "_sort_model_potential", "_crm_row_potential_eur", "_sort_input_row"],
        ascending=[True, False, False, False, True],
        kind="mergesort",
    )
    collapsed = ordered.drop_duplicates(subset=["_customer_key"], keep="first").copy()

    group_sizes = frame.groupby("_customer_key").size().rename("crm_source_row_count")
    source_ids = (
        frame.groupby("_customer_key")["_input_row_id"]
        .apply(lambda values: ", ".join(values.dropna().astype(int).astype(str).tolist()[:50]))
        .rename("crm_source_input_row_ids")
        if "_input_row_id" in frame.columns
        else pd.Series(dtype="object", name="crm_source_input_row_ids")
    )
    crm_potential = frame.groupby("_customer_key")["_crm_row_potential_eur"].sum().rename("_crm_group_potential_eur")
    collapsed = collapsed.merge(group_sizes, on="_customer_key", how="left")
    collapsed = collapsed.merge(source_ids, on="_customer_key", how="left")
    collapsed = collapsed.merge(crm_potential, on="_customer_key", how="left")

    if sales_col:
        collapsed[sales_col] = collapsed["_crm_group_potential_eur"]
    if probability_col:
        collapsed[probability_col] = 1.0

    collapsed = collapsed.drop(columns=["_customer_key", "_crm_row_potential_eur", "_crm_group_potential_eur", "_has_model", "_sort_model_potential", "_sort_input_row"], errors="ignore")
    collapsed = collapsed.sort_values(
        ["model_estimated_potential_eur", "_input_row_id"],
        ascending=[False, True],
        na_position="last",
        kind="mergesort",
    ).reset_index(drop=True)
    collapsed["customer_potential_rank"] = pd.to_numeric(collapsed["model_estimated_potential_eur"], errors="coerce").rank(
        method="first",
        ascending=False,
        na_option="bottom",
    ).astype("Int64")
    return collapsed


def _clean_product_key(value: Any) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip().lower()
    return text or None


def _build_product_lookup(grouping: pd.DataFrame) -> pd.DataFrame:
    key_columns = [column for column in ["sku", "code", "product_id", "id", "product_name", "title_fi"] if column in grouping.columns]
    rows = []
    base_cols = ["lowest_product_group_code", "lowest_product_group_name", "lowest_product_group_level"]
    for column in key_columns:
        subset = grouping[[column] + base_cols].copy()
        subset["product_key"] = subset[column].map(_clean_product_key)
        subset["product_key_source"] = column
        rows.append(subset.drop(columns=[column]))
    if not rows:
        raise ValueError("Product grouping file has no usable product key columns.")
    lookup = pd.concat(rows, ignore_index=True)
    lookup = lookup.dropna(subset=["product_key", "lowest_product_group_code"]).drop_duplicates(subset=["product_key"])
    return lookup


def build_product_group_recommendations(
    customer_potential: pd.DataFrame,
    sales: pd.DataFrame,
    accounts: pd.DataFrame,
    product_grouping: pd.DataFrame,
    *,
    max_recommendations_per_customer: int = 5,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    product_lookup = _build_product_lookup(product_grouping)
    account_frame = accounts.copy()
    account_id_col = resolve_existing_column(account_frame, ["id", "account_id", "account id"])
    business_col = resolve_existing_column(account_frame, ["business id", "business_id", "y tunnus", "y-tunnus"])
    if account_id_col is None or business_col is None:
        raise ValueError("Accounts file must contain account ID and business ID columns for sales history recommendations.")

    account_keys = account_frame[[account_id_col, business_col]].copy()
    account_keys.columns = ["account_id", "business_id"]
    account_keys["account_id"] = pd.to_numeric(account_keys["account_id"], errors="coerce")
    account_keys["business_id"] = account_keys["business_id"].map(normalize_business_id)
    account_keys = account_keys.dropna(subset=["account_id", "business_id"]).drop_duplicates(subset=["account_id"])

    sales_frame = sales.copy()
    sales_frame["account_id"] = pd.to_numeric(sales_frame["account_id"], errors="coerce")
    sales_frame["sales_eur"] = pd.to_numeric(sales_frame.get("total_value"), errors="coerce").fillna(0.0)
    sales_frame["product_key"] = pd.NA
    for column in ["sku", "code", "product_id", "id", "name"]:
        if column in sales_frame.columns:
            sales_frame["product_key"] = sales_frame["product_key"].where(
                sales_frame["product_key"].notna(),
                sales_frame[column].map(_clean_product_key),
            )
    sales_with_customer = sales_frame.merge(account_keys, on="account_id", how="left")
    sales_with_group = sales_with_customer.merge(product_lookup, on="product_key", how="left")

    grouped_sales = (
        sales_with_group.dropna(subset=["business_id", "lowest_product_group_code"])
        .groupby(["business_id", "lowest_product_group_code", "lowest_product_group_name"], as_index=False)["sales_eur"]
        .sum()
        .rename(columns={"sales_eur": "customer_sales_eur"})
    )
    total_sales = (
        grouped_sales.groupby(["lowest_product_group_code", "lowest_product_group_name"], as_index=False)["customer_sales_eur"]
        .sum()
        .rename(columns={"customer_sales_eur": "total_group_sales_eur"})
    )
    customer_totals = grouped_sales.groupby("business_id")["customer_sales_eur"].sum().rename("customer_total_sales_eur")
    grouped_sales = grouped_sales.merge(customer_totals, on="business_id", how="left")
    grouped_sales["customer_group_share"] = np.where(
        grouped_sales["customer_total_sales_eur"].gt(0),
        grouped_sales["customer_sales_eur"] / grouped_sales["customer_total_sales_eur"],
        0.0,
    )

    segment_columns = [
        "business_id",
        "company_segment",
        "model_estimated_potential_eur",
        "probability_of_growth",
        "conditional_potential_eur",
        "expected_potential_eur",
    ]
    customer_segments = customer_potential[[column for column in segment_columns if column in customer_potential.columns]].dropna(subset=["business_id"]).drop_duplicates("business_id")
    group_segment_sales = grouped_sales.merge(customer_segments[["business_id", "company_segment"]], on="business_id", how="left")
    segment_totals = group_segment_sales.groupby(["company_segment", "business_id"])["customer_sales_eur"].sum().groupby("company_segment").sum()
    segment_group = group_segment_sales.groupby(["company_segment", "lowest_product_group_code"], as_index=False)["customer_sales_eur"].sum()
    segment_group["similar_customer_group_share"] = segment_group.apply(
        lambda row: float(row["customer_sales_eur"]) / float(segment_totals.get(row["company_segment"], 0.0))
        if float(segment_totals.get(row["company_segment"], 0.0)) > 0
        else 0.0,
        axis=1,
    )

    customer_base = customer_segments.copy()
    all_groups = total_sales.copy()
    candidates = customer_base.assign(_join_key=1).merge(all_groups.assign(_join_key=1), on="_join_key").drop(columns=["_join_key"])
    candidates = candidates.merge(
        grouped_sales[["business_id", "lowest_product_group_code", "customer_sales_eur", "customer_group_share"]],
        on=["business_id", "lowest_product_group_code"],
        how="left",
    )
    candidates[["customer_sales_eur", "customer_group_share"]] = candidates[["customer_sales_eur", "customer_group_share"]].fillna(0.0)
    candidates = candidates.merge(
        segment_group[["company_segment", "lowest_product_group_code", "similar_customer_group_share"]],
        on=["company_segment", "lowest_product_group_code"],
        how="left",
    )
    global_share = total_sales["total_group_sales_eur"] / total_sales["total_group_sales_eur"].sum() if total_sales["total_group_sales_eur"].sum() else 0
    global_share_frame = total_sales[["lowest_product_group_code"]].copy()
    global_share_frame["global_group_share"] = global_share
    candidates = candidates.merge(global_share_frame, on="lowest_product_group_code", how="left")
    candidates["similar_customer_group_share"] = candidates["similar_customer_group_share"].fillna(candidates["global_group_share"]).fillna(0.0)
    candidates["white_space_gap"] = (candidates["similar_customer_group_share"] - candidates["customer_group_share"]).clip(lower=0.0)
    candidates["recommended_group_potential_eur"] = (
        pd.to_numeric(candidates["model_estimated_potential_eur"], errors="coerce").fillna(0.0) * candidates["white_space_gap"]
    )
    if "expected_potential_eur" in candidates.columns:
        candidates["recommended_group_expected_potential_eur"] = (
            pd.to_numeric(candidates["expected_potential_eur"], errors="coerce").fillna(0.0) * candidates["white_space_gap"]
        )
    else:
        candidates["recommended_group_expected_potential_eur"] = np.nan
    candidates = candidates.loc[candidates["white_space_gap"].gt(0)].copy()
    candidates = candidates.sort_values(
        ["business_id", "recommended_group_potential_eur", "total_group_sales_eur"],
        ascending=[True, False, False],
        kind="mergesort",
    )
    candidates["recommendation_rank"] = candidates.groupby("business_id").cumcount() + 1
    recommendations = candidates.loc[candidates["recommendation_rank"] <= max_recommendations_per_customer].copy()
    recommendations = recommendations.rename(
        columns={
            "lowest_product_group_code": "product_group_code",
            "lowest_product_group_name": "product_group_name",
        }
    )
    output_columns = [
        "business_id",
        "company_segment",
        "product_group_code",
        "product_group_name",
        "recommendation_rank",
        "customer_sales_eur",
        "total_group_sales_eur",
        "customer_group_share",
        "similar_customer_group_share",
        "white_space_gap",
        "probability_of_growth",
        "conditional_potential_eur",
        "expected_potential_eur",
        "recommended_group_potential_eur",
        "recommended_group_expected_potential_eur",
    ]
    missing_group_count = int(sales_with_group["lowest_product_group_code"].isna().sum())
    data_quality = pd.DataFrame(
        [
            {"metric": "sales_rows", "value": int(len(sales_with_group))},
            {"metric": "sales_rows_missing_product_group", "value": missing_group_count},
            {"metric": "product_lookup_keys", "value": int(len(product_lookup))},
            {"metric": "product_group_recommendation_rows", "value": int(len(recommendations))},
        ]
    )
    return recommendations[output_columns].reset_index(drop=True), data_quality


def validate_against_crm(customer_potential: pd.DataFrame, all_scored: pd.DataFrame | None = None) -> pd.DataFrame:
    frame = customer_potential.copy()
    sales_col = resolve_existing_column(frame, ["sales", "crm potential", "potential", "potential eur"])
    probability_col = resolve_existing_column(frame, ["probability", "prob"])
    crm_sales = pd.to_numeric(frame[sales_col], errors="coerce") if sales_col else pd.Series(np.nan, index=frame.index)
    probability = pd.to_numeric(frame[probability_col], errors="coerce") if probability_col else pd.Series(1.0, index=frame.index)
    probability = probability.where(probability.le(1.0), probability / 100.0).fillna(1.0)
    frame["crm_potential_eur"] = crm_sales.fillna(0.0) * probability
    frame["model_estimated_potential_eur"] = pd.to_numeric(frame.get("model_estimated_potential_eur"), errors="coerce")
    frame["potential_diff_eur"] = frame["model_estimated_potential_eur"] - frame["crm_potential_eur"]
    frame["potential_diff_pct"] = np.where(
        frame["crm_potential_eur"].abs().gt(0),
        frame["potential_diff_eur"] / frame["crm_potential_eur"],
        np.nan,
    )
    close_threshold = np.maximum(frame["crm_potential_eur"].abs() * 0.10, 1000.0)
    frame["validation_match_status"] = np.select(
        [
            frame["business_id"].isna(),
            frame["model_estimated_potential_eur"].isna(),
            frame["potential_diff_eur"].abs().le(close_threshold),
            frame["potential_diff_eur"].gt(close_threshold),
            frame["potential_diff_eur"].lt(-close_threshold),
        ],
        [
            "missing_business_id",
            "missing_in_model",
            "exact_or_close",
            "model_higher",
            "crm_higher",
        ],
        default="missing_in_crm",
    )
    if all_scored is not None and "is_account_customer" in all_scored.columns:
        crm_ids = set(frame["business_id"].dropna().astype(str))
        missing_model = all_scored.loc[
            all_scored["is_account_customer"].fillna(False)
            & all_scored["business_id"].notna()
            & ~all_scored["business_id"].astype(str).isin(crm_ids)
        ].copy()
        if not missing_model.empty:
            append_rows = pd.DataFrame(
                {
                    "Link": pd.NA,
                    "Name": missing_model.get("company", missing_model["business_id"]),
                    "business_id": missing_model["business_id"],
                    "rank": missing_model.get("rank"),
                    "priority": missing_model.get("priority"),
                    "company": missing_model.get("company"),
                    "score": missing_model.get("score"),
                    "model_estimated_potential_eur": missing_model.get("estimated_potential_eur"),
                    "crm_potential_eur": np.nan,
                    "potential_diff_eur": np.nan,
                    "potential_diff_pct": np.nan,
                    "validation_match_status": "missing_in_crm",
                }
            )
            frame = pd.concat([frame, append_rows], ignore_index=True, sort=False)
    return frame


def build_run_log(
    crm: pd.DataFrame,
    customer_potential: pd.DataFrame,
    validation: pd.DataFrame,
    missing_features: dict[str, int],
    product_quality: pd.DataFrame,
    artifacts: dict[str, Any],
) -> pd.DataFrame:
    matched_rows = int(customer_potential["model_estimated_potential_eur"].notna().sum())
    missing_business_rows = int(customer_potential["business_id"].isna().sum())
    rows = [
        {"metric": "input_rows", "value": int(len(crm))},
        {"metric": "output_rows", "value": int(len(customer_potential))},
        {"metric": "matched_rows", "value": matched_rows},
        {"metric": "unmatched_rows", "value": int(len(customer_potential) - matched_rows)},
        {"metric": "missing_business_id_rows", "value": missing_business_rows},
        {"metric": "reference_date", "value": str(artifacts["reference_date"].date())},
        {"metric": "model_roc_auc", "value": artifacts["metrics"].get("roc_auc")},
        {"metric": "model_average_precision", "value": artifacts["metrics"].get("average_precision")},
        {"metric": "recent_year_weight", "value": artifacts.get("recent_year_weight")},
        {"metric": "middle_year_weight", "value": artifacts.get("middle_year_weight")},
        {"metric": "oldest_year_weight", "value": artifacts.get("oldest_year_weight")},
        {"metric": "current_customer_recent_sales_weight", "value": artifacts.get("current_customer_recent_sales_weight")},
        {"metric": "recent_sales_floor_multiplier", "value": artifacts.get("recent_sales_floor_multiplier")},
    ]
    rows.extend({"metric": f"missing_feature_{key}", "value": value} for key, value in missing_features.items())
    rows.extend(product_quality.to_dict(orient="records"))
    return pd.DataFrame(rows)


def build_data_quality(crm_features: pd.DataFrame, customer_potential: pd.DataFrame, product_quality: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in customer_potential.iterrows():
        reasons = []
        if pd.isna(row.get("business_id")):
            reasons.append("missing_business_id")
        if pd.isna(row.get("model_estimated_potential_eur")):
            reasons.append("not_matched_to_model")
        if reasons:
            rows.append(
                {
                    "input_row_id": row.get("_input_row_id"),
                    "business_id": row.get("business_id"),
                    "name": row.get("Name", row.get("name")),
                    "reason": "; ".join(reasons),
                }
            )
    detail = pd.DataFrame(rows)
    if detail.empty:
        detail = pd.DataFrame(columns=["input_row_id", "business_id", "name", "reason"])
    summary = product_quality.copy()
    summary.insert(0, "input_row_id", pd.NA)
    summary["business_id"] = pd.NA
    summary["name"] = pd.NA
    summary["reason"] = summary["metric"].astype(str) + "=" + summary["value"].astype(str)
    return pd.concat([detail, summary[detail.columns]], ignore_index=True)


def remove_requested_crm_columns(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.drop(columns=[column for column in CRM_COLUMNS_F_TO_K if column in frame.columns], errors="ignore")


def write_outputs(
    customer_potential: pd.DataFrame,
    recommendations: pd.DataFrame,
    validation: pd.DataFrame,
    run_log: pd.DataFrame,
    data_quality: pd.DataFrame,
    args: argparse.Namespace,
) -> None:
    customer_potential.to_csv(args.current_customer_csv, index=False, encoding="utf-8-sig")
    recommendations.to_csv(args.recommendations_csv, index=False, encoding="utf-8-sig")
    validation.to_csv(args.validation_csv, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(args.output_xlsx, engine="openpyxl") as writer:
        customer_potential.to_excel(writer, sheet_name="customer_potential", index=False)
        recommendations.to_excel(writer, sheet_name="product_group_recommendations", index=False)
        validation.to_excel(writer, sheet_name="validation_against_crm", index=False)
        run_log.to_excel(writer, sheet_name="run_log", index=False)
        data_quality.to_excel(writer, sheet_name="data_quality", index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        workbook = load_workbook(args.output_xlsx)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200])
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 42)
        workbook.save(args.output_xlsx)
    except Exception as exc:  # pragma: no cover - formatting is best-effort.
        LOGGER.warning("Workbook formatting skipped: %s", exc)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Score Innoflame current customers with the original prospect model and product group recommendations.")
    parser.add_argument("--crm-potentials", default=str(DEFAULT_CRM_PATH))
    parser.add_argument("--product-grouping", default=str(DEFAULT_GROUPING_PATH))
    parser.add_argument("--accounts", default=str(BASE_DIR / "Account_20.05.2026_combined_with_profinder.xlsx"))
    parser.add_argument("--sales", default=str(BASE_DIR / "GoSystems_sales_26_05_2026_summarized.csv"))
    parser.add_argument("--companies", default=str(BASE_DIR / "haku_Myyntiin_ai_2026-04-23 (1).xlsx"))
    parser.add_argument("--exclude-business-ids-file", default=str(BASE_DIR / "Netvisor asiakastiedot 6-2026.xlsx"))
    parser.add_argument("--original-model", default=str(DEFAULT_ORIGINAL_MODEL))
    parser.add_argument("--v3-model", default=str(DEFAULT_V3_MODEL))
    parser.add_argument("--output-xlsx", default=str(DEFAULT_OUTPUT_XLSX))
    parser.add_argument("--current-customer-csv", default=str(DEFAULT_CURRENT_CUSTOMER_CSV))
    parser.add_argument("--recommendations-csv", default=str(DEFAULT_RECOMMENDATIONS_CSV))
    parser.add_argument("--validation-csv", default=str(DEFAULT_VALIDATION_CSV))
    parser.add_argument("--top-n-customers", type=int, default=1000)
    parser.add_argument("--lookback-days", type=int, default=365 * 3)
    parser.add_argument("--min-training-customer-annual-sales-eur", type=float, default=4000.0)
    parser.add_argument("--recent-year-weight", type=float, default=0.60)
    parser.add_argument("--middle-year-weight", type=float, default=0.30)
    parser.add_argument("--oldest-year-weight", type=float, default=0.10)
    parser.add_argument("--current-customer-recent-sales-weight", type=float, default=0.65)
    parser.add_argument("--recent-sales-floor-multiplier", type=float, default=1.00)
    parser.add_argument("--max-recommendations-per-customer", type=int, default=5)
    parser.add_argument("--random-state", type=int, default=42)
    return parser.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s %(message)s")
    args = parse_args()
    inputs = read_inputs(args)
    product_grouping, group_columns = create_lowest_product_group(inputs["product_grouping"])
    artifacts = load_model_artifacts(args, inputs)
    crm_features, matched_features = prepare_customer_features(inputs["crm"], inputs["accounts"], artifacts["all_scored"])
    customer_potential = score_current_customers(crm_features, artifacts["all_scored"])
    customer_potential = collapse_to_one_row_per_customer(customer_potential)
    recommendations, product_quality = build_product_group_recommendations(
        customer_potential,
        inputs["sales"],
        inputs["accounts"],
        product_grouping,
        max_recommendations_per_customer=args.max_recommendations_per_customer,
    )
    validation = validate_against_crm(customer_potential, artifacts["all_scored"])
    customer_potential = remove_requested_crm_columns(customer_potential)
    validation = remove_requested_crm_columns(validation)
    missing_features = {
        column: int(artifacts["modeling_df"][column].isna().sum())
        for column in artifacts["feature_columns"]
        if column in artifacts["modeling_df"].columns
    }
    product_quality = pd.concat(
        [
            product_quality,
            pd.DataFrame(
                [
                    {"metric": "product_group_level_columns_detected", "value": json.dumps(group_columns, ensure_ascii=True)},
                    {"metric": "crm_rows_matched_to_business_id", "value": int(crm_features["business_id"].notna().sum())},
                    {"metric": "crm_rows_matched_to_model", "value": int(matched_features["_has_model_score"].sum())},
                ]
            ),
        ],
        ignore_index=True,
    )
    run_log = build_run_log(inputs["crm"], customer_potential, validation, missing_features, product_quality, artifacts)
    data_quality = build_data_quality(crm_features, customer_potential, product_quality)
    write_outputs(customer_potential, recommendations, validation, run_log, data_quality, args)
    print(
        json.dumps(
            {
                "xlsx": args.output_xlsx,
                "current_customer_csv": args.current_customer_csv,
                "recommendations_csv": args.recommendations_csv,
                "validation_csv": args.validation_csv,
                "input_rows": int(len(inputs["crm"])),
                "output_rows": int(len(customer_potential)),
                "matched_rows": int(customer_potential["model_estimated_potential_eur"].notna().sum()),
                "unmatched_rows": int(customer_potential["model_estimated_potential_eur"].isna().sum()),
                "missing_business_id_rows": int(customer_potential["business_id"].isna().sum()),
                "recommendation_rows": int(len(recommendations)),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
