from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
PREVIOUS_CSV = (
    BASE
    / "two_stage_potential_model"
    / "v3_recent_weighted_current_model"
    / "innoflame_all_accounts_v3"
    / "prospect_segment_model_all_accounts_v3.csv"
)
PREVIOUS_METRICS = PREVIOUS_CSV.with_suffix(".metrics.json")
NEW_CSV = BASE / "outputs" / "innoflame_all_accounts_v3_corrected_sales" / "prospect_segment_model_all_accounts_v3_corrected_sales.csv"
NEW_METRICS = NEW_CSV.with_suffix(".metrics.json")
SALES_INPUT_AUDIT = BASE / "outputs" / "prospect_model_sales_input_invoiced_product_groups.audit.json"
OUTPUT_DIR = BASE / "outputs" / "innoflame_all_accounts_v3_corrected_sales"
COMPARISON_CSV = OUTPUT_DIR / "v3_corrected_sales_vs_previous_comparison.csv"
SUMMARY_JSON = OUTPUT_DIR / "v3_corrected_sales_vs_previous_summary.json"
SUMMARY_XLSX = OUTPUT_DIR / "v3_corrected_sales_vs_previous_summary.xlsx"


VALUE_COLUMNS = [
    "score",
    "final_value_eur",
    "expected_potential_eur",
    "ennustettu potentiaali",
    "avg_annual_sales_3y_eur",
    "recent_12m",
    "middle_12m",
    "oldest_12m",
]

BASE_COMPARE_COLUMNS = ["business_id", "company", "rank", "priority", "is_account_customer"]


def load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def summarize(frame: pd.DataFrame, label: str) -> dict[str, object]:
    non_accounts = frame.loc[~frame["is_account_customer"].astype(bool)].copy()
    final_value_column = "final_value_eur" if "final_value_eur" in frame.columns else "ennustettu potentiaali"
    return {
        "model": label,
        "rows": int(len(frame)),
        "account_customer_rows": int(frame["is_account_customer"].astype(bool).sum()),
        "non_account_rows": int(len(non_accounts)),
        "total_potential_eur": round(float(frame["ennustettu potentiaali"].sum()), 2),
        "account_customer_potential_eur": round(float(frame.loc[frame["is_account_customer"].astype(bool), "ennustettu potentiaali"].sum()), 2),
        "non_account_potential_eur": round(float(non_accounts["ennustettu potentiaali"].sum()), 2),
        "avg_score": round(float(frame["score"].mean()), 6),
        "median_final_value_eur": round(float(frame[final_value_column].median()), 2),
        "top100_potential_eur": round(float(frame.nsmallest(100, "rank")["ennustettu potentiaali"].sum()), 2),
        "top500_potential_eur": round(float(frame.nsmallest(500, "rank")["ennustettu potentiaali"].sum()), 2),
        "top1000_potential_eur": round(float(frame.nsmallest(1000, "rank")["ennustettu potentiaali"].sum()), 2),
        "top100_non_account_potential_eur": round(float(non_accounts.nsmallest(100, "rank")["ennustettu potentiaali"].sum()), 2),
    }


def top_overlap(previous: pd.DataFrame, new: pd.DataFrame, n: int, *, non_accounts_only: bool = False) -> dict[str, object]:
    prev_frame = previous
    new_frame = new
    label = f"top_{n}" + ("_non_accounts" if non_accounts_only else "")
    if non_accounts_only:
        prev_frame = prev_frame.loc[~prev_frame["is_account_customer"].astype(bool)]
        new_frame = new_frame.loc[~new_frame["is_account_customer"].astype(bool)]
    prev_ids = set(prev_frame.nsmallest(n, "rank")["business_id"].astype(str))
    new_ids = set(new_frame.nsmallest(n, "rank")["business_id"].astype(str))
    overlap = prev_ids & new_ids
    return {
        "segment": label,
        "previous_count": len(prev_ids),
        "new_count": len(new_ids),
        "overlap_count": len(overlap),
        "overlap_pct_of_new": round(100 * len(overlap) / len(new_ids), 2) if new_ids else 0.0,
        "new_entries": len(new_ids - prev_ids),
        "dropped_entries": len(prev_ids - new_ids),
    }


def main() -> None:
    previous = pd.read_csv(PREVIOUS_CSV, dtype={"business_id": str})
    new = pd.read_csv(NEW_CSV, dtype={"business_id": str})
    for frame in [previous, new]:
        if "final_value_eur" not in frame.columns and "ennustettu potentiaali" in frame.columns:
            frame["final_value_eur"] = frame["ennustettu potentiaali"]
        if "expected_potential_eur" not in frame.columns and "ennustettu potentiaali" in frame.columns:
            frame["expected_potential_eur"] = frame["ennustettu potentiaali"]
        for column in ["rank", *VALUE_COLUMNS]:
            if column in frame.columns:
                frame[column] = pd.to_numeric(frame[column], errors="coerce")
        frame["is_account_customer"] = frame["is_account_customer"].astype(str).str.lower().isin({"true", "1"})

    prev_columns = [column for column in [*BASE_COMPARE_COLUMNS, *VALUE_COLUMNS] if column in previous.columns]
    new_columns = [column for column in [*BASE_COMPARE_COLUMNS, *VALUE_COLUMNS] if column in new.columns]
    prev_subset = previous[prev_columns].copy()
    new_subset = new[new_columns].copy()
    merged = new_subset.merge(prev_subset, on="business_id", how="outer", suffixes=("_new", "_previous"), indicator=True)
    merged["rank_delta"] = merged["rank_new"] - merged["rank_previous"]
    for column in VALUE_COLUMNS:
        merged[f"{column}_delta"] = merged[f"{column}_new"] - merged[f"{column}_previous"]
    merged["company"] = merged["company_new"].fillna(merged["company_previous"])
    merged = merged.sort_values("rank_new", na_position="last")
    merged.to_csv(COMPARISON_CSV, index=False, encoding="utf-8-sig")

    previous_summary = summarize(previous, "previous_v3_2026-06-23")
    new_summary = summarize(new, "corrected_sales_v3_2026-08-24")
    summary_comparison = []
    for key in previous_summary:
        if key == "model":
            continue
        old = previous_summary[key]
        fresh = new_summary[key]
        if isinstance(old, (int, float)) and isinstance(fresh, (int, float)):
            summary_comparison.append(
                {
                    "metric": key,
                    "previous": old,
                    "new": fresh,
                    "delta": round(float(fresh) - float(old), 6),
                    "delta_pct": round(100 * (float(fresh) - float(old)) / float(old), 2) if float(old) else None,
                }
            )

    overlaps = [
        top_overlap(previous, new, 100),
        top_overlap(previous, new, 500),
        top_overlap(previous, new, 1000),
        top_overlap(previous, new, 100, non_accounts_only=True),
        top_overlap(previous, new, 500, non_accounts_only=True),
        top_overlap(previous, new, 1000, non_accounts_only=True),
    ]

    metrics = {
        "previous_metrics": load_json(PREVIOUS_METRICS),
        "new_metrics": load_json(NEW_METRICS),
        "sales_input_audit": load_json(SALES_INPUT_AUDIT),
        "summary_comparison": summary_comparison,
        "top_overlaps": overlaps,
        "largest_potential_increases": merged.loc[merged["_merge"].ne("right_only")]
        .sort_values("ennustettu potentiaali_delta", ascending=False)
        .head(25)[["business_id", "company", "rank_previous", "rank_new", "ennustettu potentiaali_previous", "ennustettu potentiaali_new", "ennustettu potentiaali_delta"]]
        .to_dict(orient="records"),
        "largest_potential_decreases": merged.loc[merged["_merge"].ne("right_only")]
        .sort_values("ennustettu potentiaali_delta", ascending=True)
        .head(25)[["business_id", "company", "rank_previous", "rank_new", "ennustettu potentiaali_previous", "ennustettu potentiaali_new", "ennustettu potentiaali_delta"]]
        .to_dict(orient="records"),
        "comparison_csv": str(COMPARISON_CSV.resolve()),
        "comparison_xlsx": str(SUMMARY_XLSX.resolve()),
    }
    SUMMARY_JSON.write_text(json.dumps(metrics, ensure_ascii=False, indent=2), encoding="utf-8")

    with pd.ExcelWriter(SUMMARY_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(summary_comparison).to_excel(writer, index=False, sheet_name="summary")
        pd.DataFrame(overlaps).to_excel(writer, index=False, sheet_name="top_overlap")
        pd.DataFrame(metrics["largest_potential_increases"]).to_excel(writer, index=False, sheet_name="largest_increases")
        pd.DataFrame(metrics["largest_potential_decreases"]).to_excel(writer, index=False, sheet_name="largest_decreases")
        merged.head(1000).to_excel(writer, index=False, sheet_name="rank_comparison_top1000")
        pd.DataFrame([metrics["sales_input_audit"]]).to_excel(writer, index=False, sheet_name="sales_input_audit")

    style_workbook(SUMMARY_XLSX)
    print(json.dumps(metrics, ensure_ascii=False, indent=2))


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                width = min(max(len(value) + 2, ws.column_dimensions[get_column_letter(cell.column)].width or 0), 50)
                ws.column_dimensions[get_column_letter(cell.column)].width = width
    wb.save(path)
    wb.close()


if __name__ == "__main__":
    main()
