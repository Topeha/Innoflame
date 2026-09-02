from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


RUN_DIR = Path("outputs/uusi_tuoteryhmittelylahde/ajo_20260826_151747")
SUMMARY_JSON = RUN_DIR / "yhteenveto.json"
BACKUPS_CSV = RUN_DIR / "vanhan_tuoteryhmittelyn_varmuuskopiot.csv"
PLAN_TXT = RUN_DIR / "suunnitelma_ja_tilanne.txt"
OUT_XLSX = RUN_DIR / "uuden_lahteen_hyodyntamissuunnitelma.xlsx"


def autosize(ws) -> None:
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        values = [str(cell.value) for cell in col_cells if cell.value is not None]
        width = min(max([len(value) for value in values] + [12]) + 2, 80)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def style_book(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in writer.book.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        autosize(ws)


def plan_rows(text: str) -> list[dict[str, str]]:
    rows = []
    current_section = ""
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line[0].isdigit() and ". " in line:
            current_section = line
            rows.append({"osio": current_section, "kohta": "", "kuvaus": line})
        elif line.startswith("- "):
            rows.append({"osio": current_section, "kohta": "tehtävä", "kuvaus": line[2:]})
        elif line in {"Huomio"}:
            current_section = line
            rows.append({"osio": current_section, "kohta": "", "kuvaus": line})
        else:
            rows.append({"osio": current_section, "kohta": "teksti", "kuvaus": line})
    return rows


def main() -> None:
    summary = json.loads(SUMMARY_JSON.read_text(encoding="utf-8"))
    plan_text = PLAN_TXT.read_text(encoding="utf-8")
    backups = pd.read_csv(BACKUPS_CSV, dtype=str, keep_default_na=False)

    zip_info = summary["source_zip"]
    overview = pd.DataFrame(
        [
            ["Ajokansio", summary["run_dir"]],
            ["Vanha tuoteryhmittely säilytetty", "Kyllä" if summary["old_grouping_preserved"] else "Ei"],
            ["Uuden lähteen polku", zip_info["path"]],
            ["Uuden lähteen tila", zip_info["status"]],
            ["Zipin koko tavua", zip_info["size_bytes"]],
            ["Tiedostoja zipissä", zip_info["member_count"]],
            ["Seuraava toimenpide", summary["next_action"]],
        ],
        columns=["Mittari", "Arvo"],
    )

    source_status = pd.DataFrame(
        [
            {
                "lähde": zip_info["path"],
                "olemassa": zip_info["exists"],
                "zip": zip_info["is_zip"],
                "koko_tavua": zip_info["size_bytes"],
                "tiedostoja_zipissä": zip_info["member_count"],
                "tila": zip_info["status"],
            }
        ]
    )

    steps = pd.DataFrame(plan_rows(plan_text))

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Yhteenveto", index=False)
        steps.to_excel(writer, sheet_name="Suunnitelma", index=False)
        backups.to_excel(writer, sheet_name="Varmuuskopiot", index=False)
        source_status.to_excel(writer, sheet_name="Uusi_lahde", index=False)
        style_book(writer)

    print(OUT_XLSX)


if __name__ == "__main__":
    main()
