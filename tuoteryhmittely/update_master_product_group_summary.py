from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path("product_master_enrichment/final_product_grouping")
FINAL_CSV = BASE / "Innoflame_tuoteryhmittely.csv"
FINAL_XLSX = BASE / "Innoflame_tuoteryhmittely.xlsx"
WORKING_XLSX = BASE / "products_product_group_tree_feedback_3level.xlsx"

OUT_DIR = Path("outputs/product_grouping_summary")
OUT_CSV = OUT_DIR / "paatotuotteet_ryhmittain_yhteenveto.csv"
OUT_XLSX = OUT_DIR / "paatotuotteet_ryhmittain_yhteenveto.xlsx"
OUT_JSON = OUT_DIR / "paatotuotteet_ryhmittain_yhteenveto.json"


def build_summary(products: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    id_col = "product_id"

    for level in [1, 2, 3]:
        group_cols: list[str] = []
        for i in range(1, level + 1):
            group_cols.extend([f"product_group_l{i}_code", f"product_group_l{i}_name"])

        grouped = (
            products.groupby(group_cols, dropna=False)[id_col]
            .nunique()
            .reset_index(name="paatotuotteita")
        )

        for _, row in grouped.iterrows():
            rows.append(
                {
                    "taso": f"L{level}",
                    "tuoteryhma_koodi": row[f"product_group_l{level}_code"],
                    "tuoteryhma": row[f"product_group_l{level}_name"],
                    "tuoteryhma_polku": " > ".join(str(row[f"product_group_l{i}_name"]) for i in range(1, level + 1)),
                    "tuoteryhma_koodipolku": " > ".join(str(row[f"product_group_l{i}_code"]) for i in range(1, level + 1)),
                    "paatotuotteita": int(row["paatotuotteita"]),
                }
            )

    return (
        pd.DataFrame(rows)
        .sort_values(["taso", "tuoteryhma_koodi", "tuoteryhma_polku"])
        .reset_index(drop=True)
    )


def autosize(ws) -> None:
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        values = [str(cell.value) for cell in column_cells[:250] if cell.value is not None]
        width = min(max([len(value) for value in values] + [12]) + 2, 60)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_summary_sheet(path: Path, summary: pd.DataFrame) -> None:
    workbook = load_workbook(path)
    insert_at = 0
    if "Yhteenveto" in workbook.sheetnames:
        insert_at = workbook.sheetnames.index("Yhteenveto")
        del workbook["Yhteenveto"]

    ws = workbook.create_sheet("Yhteenveto", insert_at)
    ws.append(list(summary.columns))
    for row in summary.itertuples(index=False, name=None):
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    autosize(ws)
    workbook.save(path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    products = pd.read_csv(FINAL_CSV, dtype=str, keep_default_na=False, low_memory=False)
    row_count = len(products)
    master_count = products["product_id"].nunique()
    if row_count != master_count:
        raise ValueError(f"Expected one row per master product, got rows={row_count}, unique product_id={master_count}")

    summary = build_summary(products)
    summary.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Yhteenveto", index=False)
        for worksheet in writer.book.worksheets:
            autosize(worksheet)

    for workbook_path in [FINAL_XLSX, WORKING_XLSX]:
        if workbook_path.exists():
            write_summary_sheet(workbook_path, summary)

    payload = {
        "source": str(FINAL_CSV.resolve()),
        "master_products": int(master_count),
        "variants_not_included": True,
        "level_group_counts": {
            level: int(summary[summary["taso"].eq(level)].shape[0])
            for level in ["L1", "L2", "L3"]
        },
        "top_l1_groups": summary[summary["taso"].eq("L1")]
        .sort_values("paatotuotteita", ascending=False)
        .head(20)
        .to_dict(orient="records"),
        "output_csv": str(OUT_CSV.resolve()),
        "output_xlsx": str(OUT_XLSX.resolve()),
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
