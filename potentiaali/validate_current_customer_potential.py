from __future__ import annotations

import argparse
import json
from pathlib import Path

import numpy as np
import pandas as pd


BASE_DIR = Path(__file__).resolve().parent


def read_outputs(args: argparse.Namespace) -> dict[str, pd.DataFrame]:
    return {
        "customer": pd.read_csv(args.customer_potential, low_memory=False),
        "recommendations": pd.read_csv(args.product_group_recommendations, low_memory=False),
        "validation": pd.read_csv(args.validation_against_crm, low_memory=False),
    }


def numeric(frame: pd.DataFrame, column: str, default: float = 0.0) -> pd.Series:
    if column not in frame.columns:
        return pd.Series(default, index=frame.index, dtype=float)
    return pd.to_numeric(frame[column], errors="coerce")


def build_technical_checks(customer: pd.DataFrame, recommendations: pd.DataFrame, validation: pd.DataFrame) -> pd.DataFrame:
    business_duplicates = (
        int(customer.loc[customer["business_id"].notna(), "business_id"].duplicated().sum())
        if "business_id" in customer.columns
        else np.nan
    )
    name_duplicates = (
        int(customer.loc[customer["business_id"].isna() & customer["_normalized_name"].notna(), "_normalized_name"].duplicated().sum())
        if {"business_id", "_normalized_name"}.issubset(customer.columns)
        else np.nan
    )
    matched_rows = int(numeric(customer, "model_estimated_potential_eur").notna().sum())
    checks = [
        {"check": "customer_potential_rows", "value": int(len(customer)), "status": "info"},
        {"check": "one_row_per_non_empty_business_id", "value": business_duplicates, "status": "pass" if business_duplicates == 0 else "review"},
        {"check": "one_row_per_missing_id_normalized_name", "value": name_duplicates, "status": "pass" if name_duplicates == 0 else "review"},
        {"check": "matched_model_rows", "value": matched_rows, "status": "info"},
        {"check": "unmatched_model_rows", "value": int(len(customer) - matched_rows), "status": "review" if len(customer) - matched_rows else "pass"},
        {"check": "missing_business_id_rows", "value": int(customer["business_id"].isna().sum()), "status": "review" if customer["business_id"].isna().any() else "pass"},
        {"check": "current_customer_true_rows", "value": int(customer.get("is_account_customer", pd.Series(False, index=customer.index)).eq(True).sum()), "status": "pass"},
        {"check": "recommendation_rows", "value": int(len(recommendations)), "status": "info"},
        {"check": "recommendation_sku_or_product_id_columns", "value": ", ".join([c for c in recommendations.columns if "sku" in c.lower() or "product_id" in c.lower()]), "status": "review" if any("sku" in c.lower() or "product_id" in c.lower() for c in recommendations.columns) else "pass"},
        {"check": "validation_rows", "value": int(len(validation)), "status": "info"},
        {"check": "duplicate_business_id_rows_in_customer_output", "value": business_duplicates, "status": "pass" if business_duplicates == 0 else "review"},
    ]
    return pd.DataFrame(checks)


def build_validation_status_summary(validation: pd.DataFrame) -> pd.DataFrame:
    summary = validation["validation_match_status"].fillna("missing_status").value_counts(dropna=False).rename_axis("validation_match_status").reset_index(name="rows")
    summary["share_of_validation_rows"] = summary["rows"] / len(validation) if len(validation) else 0.0
    return summary


def build_priority_summary(customer: pd.DataFrame) -> pd.DataFrame:
    frame = customer.copy()
    frame["model_estimated_potential_eur"] = numeric(frame, "model_estimated_potential_eur")
    frame["score"] = numeric(frame, "score")
    group_cols = ["priority"]
    summary = (
        frame.groupby(group_cols, dropna=False)
        .agg(
            rows=("business_id", "size"),
            matched_rows=("model_estimated_potential_eur", lambda values: int(values.notna().sum())),
            avg_score=("score", "mean"),
            median_potential_eur=("model_estimated_potential_eur", "median"),
            total_potential_eur=("model_estimated_potential_eur", "sum"),
        )
        .reset_index()
        .sort_values("priority", na_position="last")
    )
    return summary


def build_segment_summary(customer: pd.DataFrame) -> pd.DataFrame:
    frame = customer.copy()
    frame["model_estimated_potential_eur"] = numeric(frame, "model_estimated_potential_eur")
    frame["score"] = numeric(frame, "score")
    summary = (
        frame.groupby("company_segment", dropna=False)
        .agg(
            rows=("business_id", "size"),
            avg_score=("score", "mean"),
            median_score=("score", "median"),
            median_potential_eur=("model_estimated_potential_eur", "median"),
            total_potential_eur=("model_estimated_potential_eur", "sum"),
        )
        .reset_index()
        .sort_values(["total_potential_eur", "rows"], ascending=[False, False])
    )
    return summary


def build_distribution_summary(customer: pd.DataFrame, validation: pd.DataFrame) -> pd.DataFrame:
    rows = []
    definitions = [
        ("customer_score", customer, "score"),
        ("customer_model_estimated_potential_eur", customer, "model_estimated_potential_eur"),
        ("crm_potential_eur", validation, "crm_potential_eur"),
        ("potential_diff_eur", validation, "potential_diff_eur"),
        ("recommended_group_potential_eur", None, None),
    ]
    for metric, frame, column in definitions[:-1]:
        values = numeric(frame, column).dropna()
        rows.append(_distribution_row(metric, values))
    return pd.DataFrame(rows)


def _distribution_row(metric: str, values: pd.Series) -> dict[str, float | int | str]:
    if values.empty:
        return {"metric": metric, "count": 0}
    return {
        "metric": metric,
        "count": int(values.count()),
        "min": float(values.min()),
        "p10": float(values.quantile(0.10)),
        "p25": float(values.quantile(0.25)),
        "median": float(values.quantile(0.50)),
        "p75": float(values.quantile(0.75)),
        "p90": float(values.quantile(0.90)),
        "max": float(values.max()),
        "mean": float(values.mean()),
        "sum": float(values.sum()),
    }


def build_prior_output_comparison(customer: pd.DataFrame) -> pd.DataFrame:
    candidates = [
        BASE_DIR / "two_stage_potential_model" / "v3_recent_weighted_current_model" / "innoflame_all_accounts_v3" / "prospect_segment_model_all_accounts_v3_customers_only.csv",
        BASE_DIR / "prospect_segment_model_all_prospects.csv",
        BASE_DIR / "prospect_scoring_output.csv",
    ]
    current = customer.copy()
    current["score"] = numeric(current, "score")
    current["potential"] = numeric(current, "model_estimated_potential_eur")
    rows = [_comparison_stats("current_customer_potential.csv", current, "score", "potential")]
    for path in candidates:
        if not path.exists():
            continue
        frame = pd.read_csv(path, low_memory=False)
        score_col = "score" if "score" in frame.columns else None
        potential_col = next((c for c in ["estimated_potential_eur", "ennustettu potentiaali", "final_value_eur", "model_estimated_potential_eur"] if c in frame.columns), None)
        rows.append(_comparison_stats(str(path.relative_to(BASE_DIR)), frame, score_col, potential_col))
    return pd.DataFrame(rows)


def _comparison_stats(source: str, frame: pd.DataFrame, score_col: str | None, potential_col: str | None) -> dict[str, float | int | str | None]:
    score = numeric(frame, score_col).dropna() if score_col else pd.Series(dtype=float)
    potential = numeric(frame, potential_col).dropna() if potential_col else pd.Series(dtype=float)
    return {
        "source": source,
        "rows": int(len(frame)),
        "score_count": int(score.count()),
        "score_median": float(score.median()) if not score.empty else np.nan,
        "score_p90": float(score.quantile(0.90)) if not score.empty else np.nan,
        "potential_column": potential_col,
        "potential_count": int(potential.count()),
        "potential_median_eur": float(potential.median()) if not potential.empty else np.nan,
        "potential_p90_eur": float(potential.quantile(0.90)) if not potential.empty else np.nan,
        "potential_total_eur": float(potential.sum()) if not potential.empty else np.nan,
    }


def select_review_queues(customer: pd.DataFrame, recommendations: pd.DataFrame, validation: pd.DataFrame) -> dict[str, pd.DataFrame]:
    validation = validation.copy()
    validation["abs_potential_diff_eur"] = numeric(validation, "potential_diff_eur").abs()
    validation["crm_potential_eur"] = numeric(validation, "crm_potential_eur")
    validation["model_estimated_potential_eur"] = numeric(validation, "model_estimated_potential_eur")

    base_cols = [
        "Name",
        "company",
        "business_id",
        "priority",
        "score",
        "crm_potential_eur",
        "model_estimated_potential_eur",
        "potential_diff_eur",
        "potential_diff_pct",
        "validation_match_status",
        "positive_signals",
    ]
    base_cols = [c for c in base_cols if c in validation.columns]

    top_abs_diff = validation.loc[validation["potential_diff_eur"].notna()].sort_values("abs_potential_diff_eur", ascending=False).head(100)
    model_higher = validation.loc[validation["validation_match_status"].eq("model_higher")].sort_values("potential_diff_eur", ascending=False).head(100)
    crm_higher = validation.loc[validation["validation_match_status"].eq("crm_higher")].sort_values("potential_diff_eur", ascending=True).head(100)
    zero_crm_high_model = validation.loc[
        validation["crm_potential_eur"].fillna(0).eq(0)
        & validation["model_estimated_potential_eur"].fillna(0).gt(0)
    ].sort_values("model_estimated_potential_eur", ascending=False).head(100)

    customer_frame = customer.copy()
    customer_frame["score"] = numeric(customer_frame, "score")
    customer_frame["model_estimated_potential_eur"] = numeric(customer_frame, "model_estimated_potential_eur")
    top_a_customers = customer_frame.loc[customer_frame["priority"].eq("A")].sort_values("model_estimated_potential_eur", ascending=False).head(100)
    large_low_priority = customer_frame.loc[
        customer_frame["priority"].isin(["C", "D"])
        & customer_frame["model_estimated_potential_eur"].fillna(0).gt(customer_frame["model_estimated_potential_eur"].quantile(0.90))
    ].sort_values("model_estimated_potential_eur", ascending=False).head(100)
    missing_business = customer_frame.loc[customer_frame["business_id"].isna()].head(500)

    recommendation_frame = recommendations.copy()
    recommendation_frame["recommended_group_potential_eur"] = numeric(recommendation_frame, "recommended_group_potential_eur")
    recommendation_frame["white_space_gap"] = numeric(recommendation_frame, "white_space_gap")
    top_white_space = recommendation_frame.sort_values("recommended_group_potential_eur", ascending=False).head(100)
    suspicious_recommendations = recommendation_frame.loc[
        recommendation_frame["customer_sales_eur"].fillna(0).gt(0)
        & recommendation_frame["white_space_gap"].fillna(0).gt(0)
    ].sort_values("recommended_group_potential_eur", ascending=False).head(100)

    customer_cols = [
        "Name",
        "company",
        "business_id",
        "priority",
        "score",
        "model_estimated_potential_eur",
        "company_segment",
        "industry",
        "positive_signals",
    ]
    customer_cols = [c for c in customer_cols if c in customer.columns]

    return {
        "top_100_abs_crm_model_diff": top_abs_diff[base_cols + ["abs_potential_diff_eur"]],
        "top_100_model_higher": model_higher[base_cols],
        "top_100_crm_higher": crm_higher[base_cols],
        "top_100_zero_crm_high_model": zero_crm_high_model[base_cols],
        "top_100_a_priority_customers": top_a_customers[customer_cols],
        "large_low_priority_review": large_low_priority[customer_cols],
        "missing_business_id_review": missing_business[[c for c in ["_input_row_id", "Name", "CRM Group", "Type", "Status", "Sales", "Comment", "_normalized_name"] if c in missing_business.columns]],
        "top_100_product_group_white_space": top_white_space,
        "product_group_sanity_review": suspicious_recommendations,
    }


def build_recommendation_summary(recommendations: pd.DataFrame) -> pd.DataFrame:
    frame = recommendations.copy()
    frame["recommended_group_potential_eur"] = numeric(frame, "recommended_group_potential_eur")
    frame["white_space_gap"] = numeric(frame, "white_space_gap")
    summary = (
        frame.groupby(["product_group_code", "product_group_name"], dropna=False)
        .agg(
            recommendation_rows=("business_id", "size"),
            customers=("business_id", "nunique"),
            total_recommended_potential_eur=("recommended_group_potential_eur", "sum"),
            median_white_space_gap=("white_space_gap", "median"),
            avg_similar_customer_group_share=("similar_customer_group_share", "mean"),
            avg_customer_group_share=("customer_group_share", "mean"),
        )
        .reset_index()
        .sort_values("total_recommended_potential_eur", ascending=False)
    )
    return summary


def write_reports(reports: dict[str, pd.DataFrame], args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    for name, frame in reports.items():
        frame.to_csv(output_dir / f"{name}.csv", index=False, encoding="utf-8-sig")

    excel_path = Path(args.output_xlsx)
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for name, frame in reports.items():
            sheet_name = name[:31]
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    workbook = load_workbook(excel_path)
    fill = PatternFill("solid", fgColor="244062")
    font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
        for col in sheet.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col[:200])
            sheet.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 48)
    workbook.save(excel_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build validation audit reports for current-customer potential output.")
    parser.add_argument("--customer-potential", default=str(BASE_DIR / "current_customer_potential.csv"))
    parser.add_argument("--product-group-recommendations", default=str(BASE_DIR / "product_group_recommendations.csv"))
    parser.add_argument("--validation-against-crm", default=str(BASE_DIR / "validation_against_crm.csv"))
    parser.add_argument("--output-dir", default=str(BASE_DIR / "validation_audit_outputs"))
    parser.add_argument("--output-xlsx", default=str(BASE_DIR / "current_customer_potential_validation_audit.xlsx"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    data = read_outputs(args)
    customer = data["customer"]
    recommendations = data["recommendations"]
    validation = data["validation"]

    reports: dict[str, pd.DataFrame] = {
        "technical_checks": build_technical_checks(customer, recommendations, validation),
        "validation_status_summary": build_validation_status_summary(validation),
        "priority_summary": build_priority_summary(customer),
        "segment_summary": build_segment_summary(customer),
        "distribution_summary": build_distribution_summary(customer, validation),
        "prior_output_comparison": build_prior_output_comparison(customer),
        "product_group_summary": build_recommendation_summary(recommendations),
    }
    reports.update(select_review_queues(customer, recommendations, validation))
    write_reports(reports, args)

    print(
        json.dumps(
            {
                "output_xlsx": args.output_xlsx,
                "output_dir": args.output_dir,
                "reports": list(reports),
                "customer_rows": int(len(customer)),
                "recommendation_rows": int(len(recommendations)),
                "validation_rows": int(len(validation)),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
