from __future__ import annotations

import gzip
import json
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import openpyxl  # noqa: F401
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


DEFAULT_INPUT_PATH = Path(
    r"C:\Users\TommiHavukainen\OneDrive - Unikie Oy\Customer\Innoflame\products.json.gz"
)

PRODUCT_CONTAINER_KEYS = ("products", "items", "data", "results", "records")
DESCRIPTION_CANDIDATES = (
    "description_fi",
    "description_en",
    "description",
    "Description",
    "Descripson",
    "descripson",
    "info_description",
    "searchdata",
)
NAME_CANDIDATES = (
    "product_name",
    "name",
    "title",
    "title_fi",
    "title_en",
    "info_name",
    "info_title",
)
WAREHOUSE_CANDIDATES = (
    "warehouse_category",
    "Warehouse category",
    "Warehouse Category",
    "warehousecategory",
    "info_warehousecategory",
)
CATEGORY_LEVEL_2_CANDIDATES = (
    "category_level_2",
    "product_group",
    "category",
    "subcategory",
    "type",
    "collection",
    "info_categoryid",
)
CATEGORY_LEVEL_3_CANDIDATES = (
    "category_level_3",
    "subcategory",
    "type",
    "collection",
    "info_groupinfo",
)

NUM = r"\d+(?:[.,]\d+)?"
UNIT_DIM = r"mm|cm|m"
UNIT_WEIGHT = r"kg|g|gram|grams|kilogram|kilograms"
ILLEGAL_EXCEL_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

THREE_DIM_RE = re.compile(
    rf"(?P<a>{NUM})\s*[xX\u00d7]\s*(?P<b>{NUM})\s*[xX\u00d7]\s*(?P<c>{NUM})\s*(?P<unit>{UNIT_DIM})\b",
    re.IGNORECASE,
)
THREE_DIM_NO_UNIT_RE = re.compile(
    rf"(?P<a>{NUM})\s*[xX\u00d7]\s*(?P<b>{NUM})\s*[xX\u00d7]\s*(?P<c>{NUM})(?!\s*(?:{UNIT_DIM})\b)",
    re.IGNORECASE,
)
EXPLICIT_DIM_RE = re.compile(
    rf"\b(?P<label>width|w|leveys|length|l|pituus|depth|d|syvyys|height|h|korkeus)\b"
    rf"\s*[:=]?\s*(?P<value>{NUM})\s*(?P<unit>{UNIT_DIM})\b",
    re.IGNORECASE,
)
EXPLICIT_DIM_NO_UNIT_RE = re.compile(
    rf"\b(?P<label>width|w|leveys|length|l|pituus|depth|d|syvyys|height|h|korkeus)\b"
    rf"\s*[:=]?\s*(?P<value>{NUM})(?!\s*(?:{UNIT_DIM})\b)",
    re.IGNORECASE,
)
WEIGHT_RE = re.compile(
    rf"\b(?:(?:paino|weight|net weight)\s*:?\s*)?(?P<value>{NUM})\s*(?P<unit>{UNIT_WEIGHT})\b(?!\s*/)",
    re.IGNORECASE,
)


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def clean_illegal_excel_chars(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_EXCEL_CHARS_RE.sub("", value)
    return value


def clean_dataframe_text(df: pd.DataFrame) -> pd.DataFrame:
    return df.map(clean_illegal_excel_chars) if hasattr(df, "map") else df.applymap(clean_illegal_excel_chars)


def ascii_lower(value: Any) -> str:
    text = unicodedata.normalize("NFKD", normalize_text(value))
    return text.encode("ascii", "ignore").decode("ascii").lower()


def first_non_empty(row: pd.Series, columns: list[str]) -> Any:
    for col in columns:
        if col in row and pd.notna(row[col]) and str(row[col]).strip() != "":
            return row[col]
    return ""


def load_products(input_path: Path) -> Any:
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    try:
        with gzip.open(input_path, "rt", encoding="utf-8") as handle:
            text = handle.read()
    except OSError as exc:
        raise RuntimeError(f"Could not decompress gzip file: {exc}") from exc

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        records = []
        for line_no, line in enumerate(text.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSON at line {line_no}: {exc}") from exc
        if records:
            return records
        raise


def locate_records(data: Any) -> list[Any]:
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in PRODUCT_CONTAINER_KEYS:
            value = data.get(key)
            if isinstance(value, list):
                return value
        list_values = [(key, value) for key, value in data.items() if isinstance(value, list)]
        if len(list_values) == 1:
            return list_values[0][1]
    raise ValueError("Could not locate product records in JSON structure")


def flatten_innoflame_product(record: dict[str, Any]) -> dict[str, Any]:
    flat: dict[str, Any] = {}

    info = record.get("info")
    if isinstance(info, dict):
        flat.update(pd.json_normalize(info, sep="_").iloc[0].to_dict())

    for key, value in record.items():
        if key in {"info", "texts"}:
            continue
        if isinstance(value, list):
            flat[f"{key}_count"] = len(value)
        elif isinstance(value, dict):
            nested = pd.json_normalize(value, sep="_").iloc[0].to_dict()
            flat.update({f"{key}_{nested_key}": nested_value for nested_key, nested_value in nested.items()})
        else:
            flat[key] = value

    texts = record.get("texts")
    if isinstance(texts, list):
        for text_item in texts:
            if not isinstance(text_item, dict):
                continue
            lang = str(text_item.get("lang", "")).strip().lower()
            code = str(text_item.get("code", "")).strip().lower()
            content = text_item.get("content", "")
            if not lang or not code:
                continue
            column = f"{code}_{lang}"
            if column in flat and str(flat[column]).strip():
                flat[column] = f"{flat[column]}\n{content}"
            else:
                flat[column] = content
    return flat


def normalize_products(data: Any) -> pd.DataFrame:
    records = locate_records(data)
    if not records:
        return pd.DataFrame()
    if all(isinstance(record, dict) and ("info" in record or "texts" in record) for record in records):
        flattened = [flatten_innoflame_product(record) for record in records if isinstance(record, dict)]
        return pd.DataFrame(flattened)
    return pd.json_normalize(records, sep="_")


def find_column(df: pd.DataFrame, possible_names: tuple[str, ...] | list[str]) -> str | None:
    normalized = {str(col).lower().replace(" ", "").replace("_", ""): col for col in df.columns}
    for name in possible_names:
        key = str(name).lower().replace(" ", "").replace("_", "")
        if key in normalized:
            return normalized[key]
    return None


def extract_weight(description: Any) -> dict[str, Any]:
    text = normalize_text(description)
    if not text.strip():
        return {"weight_value": None, "weight_unit": "", "weight_g": None}

    for match in WEIGHT_RE.finditer(text):
        unit = match.group("unit").lower()
        value = parse_number(match.group("value"))
        if value is None:
            continue
        if unit in {"gram", "grams"}:
            unit = "g"
        elif unit in {"kilogram", "kilograms"}:
            unit = "kg"
        weight_g = value * 1000 if unit == "kg" else value
        return {"weight_value": value, "weight_unit": unit, "weight_g": weight_g}
    return {"weight_value": None, "weight_unit": "", "weight_g": None}


def map_dimension_label(label: str) -> str:
    label = label.lower()
    if label in {"width", "w", "leveys"}:
        return "width"
    if label in {"length", "l", "pituus"}:
        return "length"
    return "depth"


def blank_dimensions(status: str, source_text: str = "") -> dict[str, Any]:
    return {
        "width_value": None,
        "width_unit": "",
        "length_value": None,
        "length_unit": "",
        "depth_value": None,
        "depth_unit": "",
        "dimension_source_text": source_text,
        "dimension_parse_status": status,
    }


def extract_dimensions(description: Any) -> dict[str, Any]:
    text = normalize_text(description)
    if not text.strip():
        return blank_dimensions("not_found")

    explicit: dict[str, tuple[float, str, str]] = {}
    for match in EXPLICIT_DIM_RE.finditer(text):
        key = map_dimension_label(match.group("label"))
        value = parse_number(match.group("value"))
        if value is not None and key not in explicit:
            explicit[key] = (value, match.group("unit").lower(), match.group(0))
    if explicit:
        result = blank_dimensions("parsed_explicit", " | ".join(item[2] for item in explicit.values()))
        for key, (value, unit, _) in explicit.items():
            result[f"{key}_value"] = value
            result[f"{key}_unit"] = unit
        return result

    missing_unit_matches = list(EXPLICIT_DIM_NO_UNIT_RE.finditer(text))
    if missing_unit_matches:
        return blank_dimensions("value_found_unit_missing", " | ".join(match.group(0) for match in missing_unit_matches))

    match = THREE_DIM_RE.search(text)
    if match:
        unit = match.group("unit").lower()
        length = parse_number(match.group("a"))
        width = parse_number(match.group("b"))
        depth = parse_number(match.group("c"))
        return {
            "width_value": width,
            "width_unit": unit,
            "length_value": length,
            "length_unit": unit,
            "depth_value": depth,
            "depth_unit": unit,
            "dimension_source_text": match.group(0),
            "dimension_parse_status": "parsed_assumed_order",
        }

    no_unit_match = THREE_DIM_NO_UNIT_RE.search(text)
    if no_unit_match:
        return blank_dimensions("value_found_unit_missing", no_unit_match.group(0))

    return blank_dimensions("not_found")


CLASSIFICATION_RULES: list[tuple[str, str, tuple[str, ...]]] = [
    ("Sisustus", "Kynttilät", (r"kynttil", r"candle")),
    ("Kattaus", "Servetit", (r"serviet", r"napkin")),
    ("Kattaus", "Astiat", (r"astia", r"lautanen", r"kulho", r"muki", r"lasi", r"cup", r"plate", r"bowl")),
    ("Koti", "Keittiö", (r"keitti", r"paist", r"veitsi", r"leikkuulauta", r"kannu", r"tarjotin", r"kitchen")),
    ("Koti", "Tekstiilit", (r"pyyhe", r"huopa", r"viltti", r"blanket", r"towel", r"tekstiil")),
    ("Koti", "Sisustus", (r"sisustus", r"koriste", r"maljakko", r"ruukku", r"decor")),
    ("Sesonki", "Joulu", (r"joulu", r"christmas", r"xmas")),
    ("Sesonki", "Pääsiäinen", (r"paasiainen", r"pääsiäinen", r"easter")),
    ("Lahjat", "Lahjapakkaukset", (r"lahja", r"gift", r"pakkaus", r"box")),
    ("Kortit ja paperi", "Kortit", (r"kortti", r"card", r"postikortti")),
    ("Pakkaukset", "Kassit ja pussit", (r"kassi", r"pussi", r"bag", r"pouch")),
    ("Tekstiilit", "Vaatteet", (r"paita", r"shirt", r"huppari", r"hoodie", r"takki", r"jacket", r"pipo", r"cap")),
    ("Juoma-astiat", "Pullot ja mukit", (r"pullo", r"bottle", r"muki", r"mug", r"termos", r"tumbler")),
    ("Makeiset ja elintarvikkeet", "Makeiset", (r"suklaa", r"kark", r"makeis", r"candy", r"chocolate", r"pastilli")),
]


def classify_product(row: pd.Series, name_col: str | None, description_col: str | None, warehouse_col: str | None) -> dict[str, Any]:
    category_level_1 = first_non_empty(row, [warehouse_col] if warehouse_col else [])

    level_2_col = find_column(pd.DataFrame(columns=row.index), CATEGORY_LEVEL_2_CANDIDATES)
    level_3_col = find_column(pd.DataFrame(columns=row.index), CATEGORY_LEVEL_3_CANDIDATES)
    existing_level_2 = first_non_empty(row, [level_2_col] if level_2_col else [])
    existing_level_3 = first_non_empty(row, [level_3_col] if level_3_col else [])
    if existing_level_2 and existing_level_3:
        return {
            "category_level_1": category_level_1,
            "category_level_2": existing_level_2,
            "category_level_3": existing_level_3,
            "category_parse_status": "existing_category_fields",
        }

    text = " ".join(
        [
            ascii_lower(first_non_empty(row, [name_col] if name_col else [])),
            ascii_lower(first_non_empty(row, [description_col] if description_col else [])),
            ascii_lower(category_level_1),
        ]
    )
    for level_2, level_3, patterns in CLASSIFICATION_RULES:
        if any(re.search(pattern, text, re.IGNORECASE) for pattern in patterns):
            return {
                "category_level_1": category_level_1,
                "category_level_2": existing_level_2 or level_2,
                "category_level_3": existing_level_3 or level_3,
                "category_parse_status": "rule_based",
            }
    return {
        "category_level_1": category_level_1,
        "category_level_2": existing_level_2 or "",
        "category_level_3": existing_level_3 or "",
        "category_parse_status": "not_classified",
    }


def ensure_master_columns(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    id_col = find_column(result, ("product_id", "id", "info_id"))
    name_col = find_column(result, NAME_CANDIDATES)
    sku_col = find_column(result, ("sku", "code", "info_code"))
    ean_col = find_column(result, ("ean", "barcode", "gtin", "info_ean", "info_barcode"))
    warehouse_col = find_column(result, WAREHOUSE_CANDIDATES)

    result["product_id"] = result[id_col] if id_col else ""
    result["product_name"] = result[name_col] if name_col else ""
    result["sku"] = result[sku_col] if sku_col else ""
    result["ean"] = result[ean_col] if ean_col else ""
    result["warehouse_category"] = result[warehouse_col] if warehouse_col else ""
    return result


def enrich_products(df: pd.DataFrame) -> pd.DataFrame:
    result = ensure_master_columns(df)
    description_col = find_column(result, DESCRIPTION_CANDIDATES)
    name_col = find_column(result, NAME_CANDIDATES)
    warehouse_col = find_column(result, WAREHOUSE_CANDIDATES)

    weights: list[dict[str, Any]] = []
    dimensions: list[dict[str, Any]] = []
    categories: list[dict[str, Any]] = []
    for _, row in result.iterrows():
        description = first_non_empty(row, [description_col] if description_col else [])
        weights.append(extract_weight(description))
        dimensions.append(extract_dimensions(description))
        categories.append(classify_product(row, name_col, description_col, warehouse_col))

    enriched = pd.concat(
        [
            result.reset_index(drop=True),
            pd.DataFrame(weights),
            pd.DataFrame(dimensions),
            pd.DataFrame(categories),
        ],
        axis=1,
    )
    return enriched


def create_quality_report(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    has_weight = df["weight_value"].notna()
    has_dimensions = df[["width_value", "length_value", "depth_value"]].notna().any(axis=1)
    has_warehouse = df["warehouse_category"].notna() & (df["warehouse_category"].astype(str).str.strip() != "")
    has_level_2 = df["category_level_2"].notna() & (df["category_level_2"].astype(str).str.strip() != "")
    has_level_3 = df["category_level_3"].notna() & (df["category_level_3"].astype(str).str.strip() != "")
    full_category = has_warehouse & has_level_2 & has_level_3

    metrics = {
        "total_products": len(df),
        "products_with_weight": int(has_weight.sum()),
        "products_without_weight": int((~has_weight).sum()),
        "products_with_dimensions": int(has_dimensions.sum()),
        "products_without_dimensions": int((~has_dimensions).sum()),
        "products_with_warehouse_category": int(has_warehouse.sum()),
        "products_without_warehouse_category": int((~has_warehouse).sum()),
        "products_with_category_level_2": int(has_level_2.sum()),
        "products_with_category_level_3": int(has_level_3.sum()),
        "products_with_full_category_hierarchy": int(full_category.sum()),
    }
    report = pd.DataFrame({"metric": list(metrics.keys()), "value": list(metrics.values())})
    missing = {
        "missing_dimensions": df.loc[~has_dimensions],
        "missing_weight": df.loc[~has_weight],
        "missing_warehouse_category": df.loc[~has_warehouse],
        "missing_category": df.loc[~full_category],
    }
    return report, missing


def safe_sheet_name(name: str) -> str:
    return name[:31].replace("/", "_").replace("\\", "_")


def format_excel_workbook(path: Path) -> None:
    if not OPENPYXL_AVAILABLE:
        return
    workbook = openpyxl.load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in workbook.worksheets:
        if sheet.max_row < 1:
            continue
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        for col_idx in range(1, sheet.max_column + 1):
            letter = get_column_letter(col_idx)
            values = [str(sheet.cell(row=row, column=col_idx).value or "") for row in range(1, min(sheet.max_row, 200) + 1)]
            sheet.column_dimensions[letter].width = min(max(max(len(value) for value in values) + 2, 10), 60)
        if sheet.max_row >= 2 and sheet.max_column >= 1:
            table_name = re.sub(r"\W+", "", sheet.title.title())[:20] + "Table"
            table = Table(displayName=table_name, ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            try:
                sheet.add_table(table)
            except ValueError:
                pass
    workbook.save(path)


def save_outputs(df: pd.DataFrame, quality_report: pd.DataFrame, missing: dict[str, pd.DataFrame], output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    df = clean_dataframe_text(df)
    quality_report = clean_dataframe_text(quality_report)
    missing = {name: clean_dataframe_text(frame) for name, frame in missing.items()}
    paths = {
        "products_enriched_csv": output_dir / "products_enriched.csv",
        "products_enriched_xlsx": output_dir / "products_enriched.xlsx",
        "products_quality_report_xlsx": output_dir / "products_quality_report.xlsx",
        "products_missing_dimensions_csv": output_dir / "products_missing_dimensions.csv",
        "products_missing_weight_csv": output_dir / "products_missing_weight.csv",
        "products_missing_category_csv": output_dir / "products_missing_category.csv",
    }
    df.to_csv(paths["products_enriched_csv"], index=False, encoding="utf-8-sig")
    missing["missing_dimensions"].to_csv(paths["products_missing_dimensions_csv"], index=False, encoding="utf-8-sig")
    missing["missing_weight"].to_csv(paths["products_missing_weight_csv"], index=False, encoding="utf-8-sig")
    missing["missing_category"].to_csv(paths["products_missing_category_csv"], index=False, encoding="utf-8-sig")

    if OPENPYXL_AVAILABLE:
        with pd.ExcelWriter(paths["products_enriched_xlsx"], engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="products_enriched")
            quality_report.to_excel(writer, index=False, sheet_name="quality_report")
            missing["missing_dimensions"].to_excel(writer, index=False, sheet_name="missing_dimensions")
            missing["missing_weight"].to_excel(writer, index=False, sheet_name="missing_weight")
            missing["missing_category"].to_excel(writer, index=False, sheet_name="missing_category")
        format_excel_workbook(paths["products_enriched_xlsx"])

        with pd.ExcelWriter(paths["products_quality_report_xlsx"], engine="openpyxl") as writer:
            quality_report.to_excel(writer, index=False, sheet_name="quality_report")
            for name, frame in missing.items():
                frame.to_excel(writer, index=False, sheet_name=safe_sheet_name(name))
        format_excel_workbook(paths["products_quality_report_xlsx"])
    return paths


def main() -> int:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT_PATH
    output_dir = input_path.parent
    try:
        data = load_products(input_path)
        raw_df = normalize_products(data)
        enriched = enrich_products(raw_df)
        quality_report, missing = create_quality_report(enriched)
        paths = save_outputs(enriched, quality_report, missing, output_dir)
    except Exception as exc:  # noqa: BLE001 - command line script should report any top-level failure clearly.
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    metrics = dict(zip(quality_report["metric"], quality_report["value"], strict=False))
    print(f"Products processed: {metrics['total_products']}")
    print(f"Products with weight: {metrics['products_with_weight']}")
    print(f"Products with dimensions: {metrics['products_with_dimensions']}")
    print(f"Products missing weight: {metrics['products_without_weight']}")
    print(f"Products missing dimensions: {metrics['products_without_dimensions']}")
    print(f"Products with full category hierarchy: {metrics['products_with_full_category_hierarchy']}")
    print()
    print("Output files created:")
    for path in paths.values():
        if path.exists():
            print(f"- {path.name}")
    if not OPENPYXL_AVAILABLE:
        print("openpyxl is not installed; Excel files were skipped.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
