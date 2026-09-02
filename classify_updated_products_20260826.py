from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd


ROOT = Path(__file__).resolve().parent
SOURCE_JSON = ROOT / "tuoteryhmittely" / "Päivitetty tuotedata 26.8.2026" / "products.json"
OUTPUT_DIR = ROOT / "tuoteryhmittely" / "Päivitetty tuotedata 26.8.2026" / "tuoteryhmittely_ajo_20260826"
SALES_GROUPING_CSV = ROOT / "outputs" / "Innoflame_merged_sales_csv_source_with_L1_L2_L3_delivery_handling.csv"
OLD_MASTER_XLSX = ROOT / "tuoteryhmittely" / "Innoflame_tuoteryhmittely_lahdedata" / "01_tuotemaster" / "products_table_view.xlsx"

BACKUP_FILES = [
    ROOT / "outputs" / "Innoflame_merged_sales_csv_source_with_L1_L2_L3_delivery_handling.xlsx",
    ROOT / "outputs" / "Innoflame_merged_sales_csv_source_full_audit.xlsx",
    ROOT / "outputs" / "Innoflame_unmatched_product_groups_by_name.xlsx",
    ROOT / "outputs" / "Innoflame_tuoteryhmittely_kooste.pptx",
    ROOT / "tuoteryhmittely" / "Innoflame_tuoteryhmittely.pptx",
]


L1_ORDER = [
    "Vaatteet",
    "Laukut ja matkatavarat",
    "Koti ja keittiö",
    "Promootio- ja tapahtumatuotteet",
    "Elektroniikka",
    "Vapaa-aika",
    "Elintarvikkeet",
    "Työkalut ja turvallisuus",
    "Hyvinvointi",
    "Korut, kellot ja aurinkolasit",
    "Lahjakortit ja hyväntekeväisyys",
    "Decalit ja merkkaustuotteet",
    "Toimistotuotteet",
    "Muut / tarkistettavat",
]


TEXT_RULES: list[tuple[str, str, str, str, str]] = [
    ("text_rule_decal_transfer", r"\b(siirtokuva|transfer logo|tekstiilimerk|kangasmerk|haalarimerk|decal)\b", "Decalit ja merkkaustuotteet", "Decalit", "Siirtokuvat ja tekstiilimerkit"),
    ("text_rule_decal_sticker", r"\b(tarra|sticker|etiketti|label|tuotelappu|product label)\b", "Decalit ja merkkaustuotteet", "Decalit", "Tarrat, etiketit ja tuotelaput"),
    ("text_rule_lanyard", r"\b(kaulanauha|lanyard)\b", "Promootio- ja tapahtumatuotteet", "Jakotuotteet", "Kaulanauhat"),
    ("text_rule_keyring", r"\b(avaimenper|keyring|keychain)\b", "Promootio- ja tapahtumatuotteet", "Jakotuotteet", "Avaimenperät"),
    ("text_rule_reflector_vest", r"\b(heijastinliivi|reflective vest|safety vest)\b", "Työkalut ja turvallisuus", "Suojaimet", "Heijastinliivit"),
    ("text_rule_reflector", r"\b(heijastin|reflector)\b", "Promootio- ja tapahtumatuotteet", "Jakotuotteet", "Heijastimet"),
    ("text_rule_badges", r"\b(pinssi|pin |rintanappi|badge)\b", "Promootio- ja tapahtumatuotteet", "Jakotuotteet", "Pinssit ja rintanapit"),
    ("text_rule_tube_scarf", r"\b(putkihuivi|tuubihuivi|buff|neck tube)\b", "Promootio- ja tapahtumatuotteet", "Jakotuotteet", "Putkihuivit"),
    ("text_rule_pens", r"\b(kynä|kuulakynä|lyijykynä|pen|pencil)\b", "Promootio- ja tapahtumatuotteet", "Jakotuotteet", "Kynät"),
    ("text_rule_event", r"\b(lippu|flag|banderolli|banner|roll[\s-]?up|messusein|beachflag|tapahtuma)\b", "Promootio- ja tapahtumatuotteet", "Tapahtumatuotteet", "Liput, banderollit ja messutuotteet"),
    ("text_rule_event_other", r"\b(ilmapallo|balloon|ranneke|wristband|messu)\b", "Promootio- ja tapahtumatuotteet", "Tapahtumatuotteet", "Muut tapahtumatuotteet"),
    ("text_rule_gift_card", r"\b(lahjakortti|gift card|vapaalippu|voucher)\b", "Lahjakortit ja hyväntekeväisyys", "Lahjakortit", "Lahjakortit ja pääsyliput"),
    ("text_rule_charity", r"\b(hyväntekeväisyys|charity|lahjoitus|donation)\b", "Lahjakortit ja hyväntekeväisyys", "Hyväntekeväisyys", "Hyväntekeväisyys"),
    ("text_rule_caps", r"\b(lippalakki|cap|keps|trucker|snapback)\b", "Vaatteet", "Asusteet", "Lippalakit"),
    ("text_rule_beanies", r"\b(pipo|beanie)\b", "Vaatteet", "Asusteet", "Pipot"),
    ("text_rule_gloves", r"\b(käsine|hanska|glove|mitten|rukkanen)\b", "Vaatteet", "Asusteet", "Käsineet"),
    ("text_rule_socks", r"\b(sukka|socks?)\b", "Vaatteet", "Asusteet", "Sukat"),
    ("text_rule_shoes", r"\b(kenkä|jalkine|shoe|sneaker|boot)\b", "Vaatteet", "Asusteet", "Jalkineet"),
    ("text_rule_belts", r"\b(vyö|belt)\b", "Vaatteet", "Asusteet", "Vyöt"),
    ("text_rule_scarves", r"\b(kaulahuivi|huivi|scarf|solmio|tie)\b", "Vaatteet", "Asusteet", "Huivit"),
    ("text_rule_tshirts", r"\b(t-?paita|tshirt|t-shirt|tee\b)\b", "Vaatteet", "Paidat ja yläosat", "T-paidat"),
    ("text_rule_polo", r"\b(polo|pikee)\b", "Vaatteet", "Paidat ja yläosat", "Pikeepaidat"),
    ("text_rule_hoodies", r"\b(huppari|hoodie|college|sweatshirt)\b", "Vaatteet", "Paidat ja yläosat", "Hupparit ja colleget"),
    ("text_rule_shirts", r"\b(kauluspaita|shirt|paita)\b", "Vaatteet", "Paidat ja yläosat", "Kauluspaidat ja paidat"),
    ("text_rule_jackets", r"\b(takki|jacket|parka|softshell|fleece|liivi|vest)\b", "Vaatteet", "Takit ja liivit", "Takit ja liivit"),
    ("text_rule_trousers", r"\b(housu|pants|trousers|shortsit|shorts)\b", "Vaatteet", "Housut ja alaosat", "Housut ja alaosat"),
    ("text_rule_workwear", r"\b(työvaate|workwear|hi-?vis|atex)\b", "Vaatteet", "Työvaatteet", "Muut työvaatteet"),
    ("text_rule_mugs", r"\b(muki|mug)\b", "Koti ja keittiö", "Juoma-astiat", "Mukit"),
    ("text_rule_bottles", r"\b(juomapullo|termospullo|pullo|bottle|thermos|termos|tumbler)\b", "Koti ja keittiö", "Juoma-astiat", "Juomapullot ja termospullot"),
    ("text_rule_glasses", r"\b(lasi|glass)\b", "Koti ja keittiö", "Juoma-astiat", "Lasit"),
    ("text_rule_grill", r"\b(grilli|grill|savustin|smoker)\b", "Koti ja keittiö", "Keittiötuotteet", "Grillit ja savustimet"),
    ("text_rule_kitchen", r"\b(keittiö|veitsi|knife|leikkuulauta|cutting board|tarjoilu|serving)\b", "Koti ja keittiö", "Keittiötuotteet", "Keittiövälineet"),
    ("text_rule_towels", r"\b(pyyhe|towel|laudeliina)\b", "Koti ja keittiö", "Kodintekstiilit", "Pyyhkeet ja laudeliinat"),
    ("text_rule_blankets", r"\b(huopa|blanket|peitto|lakana)\b", "Koti ja keittiö", "Kodintekstiilit", "Peitot, huovat ja lakanat"),
    ("text_rule_vase", r"\b(maljakko|vase|ruukku|pot)\b", "Koti ja keittiö", "Sisustus", "Maljakot ja ruukut"),
    ("text_rule_interior", r"\b(kynttilä|candle|sisustus|decor)\b", "Koti ja keittiö", "Sisustus", "Sisustustuotteet"),
    ("text_rule_backpacks", r"\b(reppu|backpack|selkäreppu)\b", "Laukut ja matkatavarat", "Reput", "Reput"),
    ("text_rule_cooler_bags", r"\b(kylmälaukku|cooler bag|cooler)\b", "Laukut ja matkatavarat", "Kylmälaukut", "Kylmälaukut"),
    ("text_rule_bags", r"\b(kassi|tote|shopper|bag)\b", "Laukut ja matkatavarat", "Kassit", "Kassit"),
    ("text_rule_luggage", r"\b(matkalaukku|trolley|luggage|travel bag)\b", "Laukut ja matkatavarat", "Matkatavarat", "Matkatavarat"),
    ("text_rule_jewelry", r"\b(koru|kaulakoru|rannekoru|korvakoru|kello|watch|jewelry)\b", "Korut, kellot ja aurinkolasit", "Korut ja kellot", "Korut ja kellot"),
    ("text_rule_glasses_wearable", r"\b(aurinkolasi|sunglasses|silmälasi)\b", "Korut, kellot ja aurinkolasit", "Silmälasit ja aurinkolasit", "Silmälasit ja aurinkolasit"),
    ("text_rule_candy", r"\b(suklaa|karkki|makeinen|candy|sweet)\b", "Elintarvikkeet", "Makeiset", "Makeiset"),
    ("text_rule_drinks", r"\b(kahvi|tee|tea|juoma|drink)\b", "Elintarvikkeet", "Juomat", "Juomat"),
    ("text_rule_food", r"\b(hunaja|food|snack|mauste|spice|öljy|oil)\b", "Elintarvikkeet", "Elintarvikkeet", "Muut elintarvikkeet"),
    ("text_rule_headphones", r"\b(kuuloke|headphone|earbud)\b", "Elektroniikka", "Audio", "Kuulokkeet"),
    ("text_rule_speakers", r"\b(kaiutin|speaker)\b", "Elektroniikka", "Audio", "Kaiuttimet"),
    ("text_rule_powerbank", r"\b(varavirtalähde|powerbank|power bank)\b", "Elektroniikka", "Virta ja lataus", "Varavirtalähteet"),
    ("text_rule_charger", r"\b(laturi|kaapeli|charger|cable|usb)\b", "Elektroniikka", "Virta ja lataus", "Laturit ja kaapelit"),
    ("text_rule_lights", r"\b(lamppu|valaisin|light|torch|taskulamppu)\b", "Elektroniikka", "Valaisimet", "Lamput ja valaisimet"),
    ("text_rule_umbrella", r"\b(sateenvarjo|umbrella)\b", "Vapaa-aika", "Ulkoilu", "Sateenvarjot"),
    ("text_rule_golf", r"\b(golf)\b", "Vapaa-aika", "Urheilu", "Golf"),
    ("text_rule_games", r"\b(peli|lelu|toy|game|yo-?yo|frisbee)\b", "Vapaa-aika", "Pelit ja lelut", "Lelut ja pelit"),
    ("text_rule_sunscreen", r"\b(aurinkorasva|sunscreen|sun cream)\b", "Hyvinvointi", "Kosmetiikka", "Aurinkotuotteet"),
    ("text_rule_cosmetics", r"\b(kosmetiikka|cosmetic|saippua|soap|voide|cream)\b", "Hyvinvointi", "Kosmetiikka", "Muut kosmetiikkatuotteet"),
    ("text_rule_ppe", r"\b(suojain|maski|kasvomaski|mask|kypärä|helmet|suojalasit|safety glasses)\b", "Työkalut ja turvallisuus", "Suojaimet", "Suojaimet"),
    ("text_rule_tools", r"\b(työkalu|tool|monitoimityökalu|multi-?tool|mittanauha|measure)\b", "Työkalut ja turvallisuus", "Työkalut", "Työkalut"),
    ("text_rule_notebooks", r"\b(muistikirja|notebook|vihko|kalenteri|calendar)\b", "Toimistotuotteet", "Muistikirjat ja kalenterit", "Muistikirjat ja kalenterit"),
    ("text_rule_office", r"\b(toimisto|office|paperi|paper|kortti|card)\b", "Toimistotuotteet", "Toimistotarvikkeet", "Muut toimistotuotteet"),
]

PRIORITY_RULES = {
    "text_rule_decal_transfer",
    "text_rule_decal_sticker",
    "text_rule_lanyard",
    "text_rule_keyring",
    "text_rule_reflector_vest",
    "text_rule_reflector",
    "text_rule_badges",
    "text_rule_tube_scarf",
    "text_rule_gift_card",
    "text_rule_charity",
}


def normalize_code(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.upper()


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def text_blob(product: dict[str, Any]) -> str:
    info = product.get("info") or {}
    parts = [clean(info.get("code")), clean(info.get("searchdata"))]
    for text in product.get("texts") or []:
        parts.append(clean(text.get("content")))
    return " ".join(part for part in parts if part)


def norm_for_match(value: str) -> str:
    value = value.lower()
    replacements = {
        "ä": "a",
        "ö": "o",
        "å": "a",
        "é": "e",
        "è": "e",
        "á": "a",
        "ü": "u",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return value


def fix_known_encoding(value: str) -> str:
    replacements = {
        "keitti�": "keittiö",
        "Keitti�": "Keittiö",
        "hyv�ntekev�isyys": "hyväntekeväisyys",
        "Hyv�ntekev�isyys": "Hyväntekeväisyys",
        "Ty�": "Työ",
        "ty�": "työ",
        "k�sittely": "käsittely",
        "K�sittely": "Käsittely",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return value


def extract_texts(product: dict[str, Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for item in product.get("texts") or []:
        lang = clean(item.get("lang")).lower()
        code = clean(item.get("code")).lower()
        content = clean(item.get("content"))
        if lang and code and content:
            out[f"{code}_{lang}"] = content
    return out


def read_products() -> list[dict[str, Any]]:
    with SOURCE_JSON.open("r", encoding="utf-8-sig") as handle:
        data = json.load(handle)
    if not isinstance(data, list):
        raise ValueError("products.json ei ole listamuotoinen.")
    return data


def build_sales_grouping_map() -> tuple[dict[str, tuple[str, str, str]], dict[str, int]]:
    if not SALES_GROUPING_CSV.exists():
        return {}, {"sales_mapping_rows": 0, "sales_mapping_keys": 0}

    usecols = [
        "productcode",
        "optioncode",
        "product_group_l1_name",
        "product_group_l2_name",
        "product_group_l3_name",
    ]
    counters: dict[str, Counter[tuple[str, str, str]]] = defaultdict(Counter)
    row_count = 0
    for chunk in pd.read_csv(
        SALES_GROUPING_CSV,
        usecols=usecols,
        chunksize=100_000,
        dtype=str,
        keep_default_na=False,
        low_memory=False,
    ):
        row_count += len(chunk)
        for record in chunk.to_dict("records"):
            group = (
                clean(record.get("product_group_l1_name")),
                clean(record.get("product_group_l2_name")),
                clean(record.get("product_group_l3_name")),
            )
            if not all(group) or group[0] == "Muut / tarkistettavat":
                continue
            for field in ("productcode", "optioncode"):
                key = normalize_code(record.get(field))
                if key:
                    counters[key][group] += 1

    mapping = {key: counts.most_common(1)[0][0] for key, counts in counters.items()}
    return mapping, {"sales_mapping_rows": row_count, "sales_mapping_keys": len(mapping)}


def classify_by_rules(product: dict[str, Any]) -> tuple[str, str, str, str, str]:
    normalized = norm_for_match(text_blob(product))
    for source, pattern, l1, l2, l3 in TEXT_RULES:
        normalized_pattern = norm_for_match(pattern)
        if re.search(normalized_pattern, normalized, flags=re.IGNORECASE):
            return l1, l2, l3, source, "medium"
    return (
        "Muut / tarkistettavat",
        "Muut / tarkistettavat",
        "Tarkistettavat",
        "manual_review_no_rule_match",
        "review",
    )


def classify_by_priority_rules(product: dict[str, Any]) -> tuple[str, str, str, str, str] | None:
    normalized = norm_for_match(text_blob(product))
    for source, pattern, l1, l2, l3 in TEXT_RULES:
        if source not in PRIORITY_RULES:
            continue
        normalized_pattern = norm_for_match(pattern)
        if re.search(normalized_pattern, normalized, flags=re.IGNORECASE):
            return l1, l2, l3, f"{source}_priority_override", "high"
    return None


def canonicalize_group(group: tuple[str, str, str]) -> tuple[str, str, str]:
    l1, l2, l3 = (fix_known_encoding(clean(part)) for part in group)

    if l1 == "Toimisto, painotuotteet ja pakkaukset":
        return "Toimistotuotteet", l2 or "Toimistotarvikkeet", l3 or "Muut toimistotuotteet"
    if l1 == "Makeiset ja elintarvikkeet":
        return "Elintarvikkeet", l2 or "Elintarvikkeet", l3 or "Muut elintarvikkeet"
    if l1 == "Toimitus- ja käsittelykulu":
        return "Muut / tarkistettavat", "Toimitus- ja käsittelykulut", "Toimitus- ja käsittelykulut"
    if l1 == "Asusteet":
        return "Vaatteet", "Asusteet", l3 or l2 or "Muut asusteet"
    if l1 == "Turvallisuus":
        return "Työkalut ja turvallisuus", "Suojaimet", l3 or l2 or "Suojaimet"
    if l1 == "Lahjat ja sesonkituotteet":
        if re.search(r"lahjakort|gift card|hyväntekeväisyys|charity", norm_for_match(f"{l2} {l3}")):
            return "Lahjakortit ja hyväntekeväisyys", l2 or "Lahjakortit", l3 or "Lahjakortit ja hyväntekeväisyys"
        return "Muut / tarkistettavat", "Muut / tarkistettavat", "Tarkistettavat"
    return l1, l2, l3


def enforce_min_l3_size(
    groups: list[tuple[str, str, str, str, str]],
    minimum: int = 5,
) -> list[tuple[str, str, str, str, str]]:
    counts = Counter((l1, l2, l3) for l1, l2, l3, _source, _confidence in groups)
    adjusted: list[tuple[str, str, str, str, str]] = []
    for l1, l2, l3, source, confidence in groups:
        if counts[(l1, l2, l3)] >= minimum:
            adjusted.append((l1, l2, l3, source, confidence))
            continue
        adjusted.append(
            (
                "Muut / tarkistettavat",
                "Muut / tarkistettavat",
                "Tarkistettavat",
                f"{source}_merged_small_l3",
                "review",
            )
        )
    return adjusted


def build_code_maps(paths: list[tuple[str, str, str]]) -> dict[tuple[str, str, str], tuple[str, str, str]]:
    l1_code = {name: f"{idx:02d}" for idx, name in enumerate(L1_ORDER, start=1)}
    l2_names: dict[str, list[str]] = defaultdict(list)
    l3_names: dict[tuple[str, str], list[str]] = defaultdict(list)
    for l1, l2, l3 in paths:
        if l2 not in l2_names[l1]:
            l2_names[l1].append(l2)
        if l3 not in l3_names[(l1, l2)]:
            l3_names[(l1, l2)].append(l3)

    for names in l2_names.values():
        names.sort()
    for names in l3_names.values():
        names.sort()

    codes: dict[tuple[str, str, str], tuple[str, str, str]] = {}
    for l1, l2, l3 in sorted(set(paths)):
        base = l1_code.get(l1, "99")
        l2_idx = l2_names[l1].index(l2) + 1
        l3_idx = l3_names[(l1, l2)].index(l3) + 1
        codes[(l1, l2, l3)] = (base, f"{base}.{l2_idx:02d}", f"{base}.{l2_idx:02d}.{l3_idx:02d}")
    return codes


def preserve_old_outputs() -> list[dict[str, Any]]:
    backup_dir = OUTPUT_DIR / "00_vanha_tuoteryhmittely_sailytetty"
    backup_dir.mkdir(parents=True, exist_ok=True)
    rows = []
    for source in BACKUP_FILES:
        row = {
            "source_path": str(source),
            "copied": False,
            "backup_path": "",
            "size_bytes": None,
        }
        if source.exists():
            target = backup_dir / source.name
            shutil.copy2(source, target)
            row.update({"copied": True, "backup_path": str(target), "size_bytes": target.stat().st_size})
        rows.append(row)
    return rows


def old_master_codes() -> set[str]:
    if not OLD_MASTER_XLSX.exists():
        return set()
    df = pd.read_excel(OLD_MASTER_XLSX, sheet_name="products", usecols=["code"], dtype=str)
    return {normalize_code(value) for value in df["code"].dropna()}


def make_workbook(products_df: pd.DataFrame, backup_df: pd.DataFrame, audit: dict[str, Any]) -> Path:
    output_xlsx = OUTPUT_DIR / "Innoflame_paivitetty_tuoteryhmittely_20260826.xlsx"
    output_csv = OUTPUT_DIR / "Innoflame_paivitetty_tuoteryhmittely_20260826.csv"
    products_df.to_csv(output_csv, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL)

    l1_summary = (
        products_df.groupby(["product_group_l1_code", "product_group_l1_name"], dropna=False)
        .agg(paatuotteet=("product_id", "count"), variantit=("variant_count", "sum"))
        .reset_index()
        .sort_values(["product_group_l1_code", "product_group_l1_name"])
    )
    l2_summary = (
        products_df.groupby(["product_group_l1_code", "product_group_l1_name", "product_group_l2_code", "product_group_l2_name"], dropna=False)
        .agg(paatuotteet=("product_id", "count"), variantit=("variant_count", "sum"))
        .reset_index()
        .sort_values(["product_group_l1_code", "product_group_l2_code"])
    )
    l3_summary = (
        products_df.groupby(
            [
                "product_group_l1_code",
                "product_group_l1_name",
                "product_group_l2_code",
                "product_group_l2_name",
                "product_group_l3_code",
                "product_group_l3_name",
            ],
            dropna=False,
        )
        .agg(paatuotteet=("product_id", "count"), variantit=("variant_count", "sum"))
        .reset_index()
        .sort_values(["product_group_l1_code", "product_group_l2_code", "product_group_l3_code"])
    )
    method_summary = (
        products_df.groupby(["product_group_source", "confidence"], dropna=False)
        .agg(paatuotteet=("product_id", "count"), variantit=("variant_count", "sum"))
        .reset_index()
        .sort_values(["paatuotteet"], ascending=False)
    )
    review_df = products_df[products_df["confidence"].eq("review")].copy()
    small_l3 = l3_summary[l3_summary["paatuotteet"] < 5].copy()

    audit_df = pd.DataFrame(
        [
            {"mittari": "paatotteet_uudessa_lahteessa", "arvo": audit["products_total"]},
            {"mittari": "variantit_uudessa_lahteessa", "arvo": audit["variants_total"]},
            {"mittari": "vanhassa_masterissa_loytyneet_koodit", "arvo": audit["old_master_codes"]},
            {"mittari": "uudet_tai_vanhasta_puuttuvat_koodit", "arvo": audit["new_or_missing_from_old_master"]},
            {"mittari": "kaytetty_warehousecategory_ohjaavana", "arvo": "Ei"},
            {"mittari": "myyntiryhmittelyrivit_luettu", "arvo": audit.get("sales_mapping_rows", 0)},
            {"mittari": "myyntiryhmittelykoodit_kaytossa", "arvo": audit.get("sales_mapping_keys", 0)},
            {"mittari": "l1_ryhmia", "arvo": l1_summary["product_group_l1_name"].nunique()},
            {"mittari": "l2_ryhmia", "arvo": l2_summary["product_group_l2_code"].nunique()},
            {"mittari": "l3_ryhmia", "arvo": l3_summary["product_group_l3_code"].nunique()},
            {"mittari": "tarkistettavat_paatuotteet", "arvo": len(review_df)},
            {"mittari": "alle_5_paatuotteen_l3_ryhmat", "arvo": len(small_l3)},
        ]
    )

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        audit_df.to_excel(writer, sheet_name="Yhteenveto", index=False)
        l1_summary.to_excel(writer, sheet_name="Taso1", index=False)
        l2_summary.to_excel(writer, sheet_name="Taso2", index=False)
        l3_summary.to_excel(writer, sheet_name="Taso3", index=False)
        method_summary.to_excel(writer, sheet_name="Luokittelutavat", index=False)
        small_l3.to_excel(writer, sheet_name="Pienet_ryhmat", index=False)
        review_df.to_excel(writer, sheet_name="Tarkistettavat", index=False)
        products_df.to_excel(writer, sheet_name="Tuotteet", index=False)
        backup_df.to_excel(writer, sheet_name="Vanha_sailytetty", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_xlsx)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            sample_values = [clean(cell.value) for cell in list(column_cells)[:200]]
            max_len = max([len(value) for value in sample_values] + [len(clean(column_cells[0].value))])
            width = min(max(max_len + 2, 10), 46)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
    wb.save(output_xlsx)
    return output_xlsx


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    backup_rows = preserve_old_outputs()
    products = read_products()
    sales_map, sales_audit = build_sales_grouping_map()
    old_codes = old_master_codes()

    rows: list[dict[str, Any]] = []
    classified_paths: list[tuple[str, str, str]] = []
    interim_groups: list[tuple[str, str, str, str, str]] = []

    for product in products:
        info = product.get("info") or {}
        texts = extract_texts(product)
        options = product.get("options") or []
        product_code = normalize_code(info.get("code"))
        option_codes = [normalize_code((option.get("info") or {}).get("code")) for option in options]
        option_codes = [code for code in option_codes if code]

        priority_group = classify_by_priority_rules(product)
        if priority_group:
            l1, l2, l3, source, confidence = priority_group
        elif product_code in sales_map:
            l1, l2, l3 = canonicalize_group(sales_map[product_code])
            source = "existing_sales_productcode"
            confidence = "high"
        else:
            option_matches = [canonicalize_group(sales_map[code]) for code in option_codes if code in sales_map]
            if option_matches:
                l1, l2, l3 = Counter(option_matches).most_common(1)[0][0]
                source = "existing_sales_optioncode"
                confidence = "high"
            else:
                l1, l2, l3, source, confidence = classify_by_rules(product)

        interim_groups.append((l1, l2, l3, source, confidence))
        classified_paths.append((l1, l2, l3))

    original_interim_groups = interim_groups.copy()
    interim_groups = enforce_min_l3_size(interim_groups, minimum=5)
    classified_paths = [(l1, l2, l3) for l1, l2, l3, _source, _confidence in interim_groups]
    code_map = build_code_maps(classified_paths)

    for product, group, original_group in zip(products, interim_groups, original_interim_groups, strict=True):
        info = product.get("info") or {}
        texts = extract_texts(product)
        options = product.get("options") or []
        product_code = normalize_code(info.get("code"))
        option_codes = [normalize_code((option.get("info") or {}).get("code")) for option in options]
        option_codes = [code for code in option_codes if code]
        active_option_count = sum(1 for option in options if (option.get("info") or {}).get("status") == 1)
        l1, l2, l3, source, confidence = group
        original_l1, original_l2, original_l3, _original_source, _original_confidence = original_group
        l1_code, l2_code, l3_code = code_map[(l1, l2, l3)]
        rows.append(
            {
                "product_id": info.get("id"),
                "code": clean(info.get("code")),
                "title_fi": texts.get("title_fi", ""),
                "description_fi": texts.get("description_fi", ""),
                "title_en": texts.get("title_en", ""),
                "description_en": texts.get("description_en", ""),
                "searchdata": clean(info.get("searchdata")),
                "status": info.get("status"),
                "discontinued": info.get("discontinued"),
                "provider_id": info.get("providerid"),
                "brand_id": info.get("brandid"),
                "price": info.get("price"),
                "price_vat": info.get("price_vat"),
                "buyprice": info.get("buyprice"),
                "warehousecategory_audit_only": info.get("warehousecategory"),
                "inventorycategory_audit_only": info.get("inventorycategory"),
                "variant_count": len(options),
                "active_variant_count": active_option_count,
                "first_option_code": option_codes[0] if option_codes else "",
                "option_codes_sample": ", ".join(option_codes[:10]),
                "is_new_or_missing_from_old_master": product_code not in old_codes,
                "product_group_l1_code": l1_code,
                "product_group_l1_name": l1,
                "product_group_l2_code": l2_code,
                "product_group_l2_name": l2,
                "product_group_l3_code": l3_code,
                "product_group_l3_name": l3,
                "original_suggested_l1_name": original_l1,
                "original_suggested_l2_name": original_l2,
                "original_suggested_l3_name": original_l3,
                "product_group_source": source,
                "confidence": confidence,
            }
        )

    products_df = pd.DataFrame(rows)
    backup_df = pd.DataFrame(backup_rows)
    audit = {
        "source_json": str(SOURCE_JSON),
        "output_dir": str(OUTPUT_DIR),
        "products_total": len(products_df),
        "variants_total": int(products_df["variant_count"].sum()),
        "old_master_codes": len(old_codes),
        "new_or_missing_from_old_master": int(products_df["is_new_or_missing_from_old_master"].sum()),
        **sales_audit,
    }

    output_xlsx = make_workbook(products_df, backup_df, audit)
    audit_path = OUTPUT_DIR / "Innoflame_paivitetty_tuoteryhmittely_20260826_audit.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps({**audit, "output_xlsx": str(output_xlsx)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
