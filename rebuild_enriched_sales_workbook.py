import csv
import json
import os
import shutil
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = r"C:\Users\TommiHavukainen\OneDrive - Unikie Oy\Customer\Innoflame"
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
CSV_PATH = os.path.join(
    OUTPUT_DIR, "Innoflame_merged_sales_csv_source_with_L1_L2_L3_delivery_handling.csv"
)
XLSX_PATH = os.path.join(
    OUTPUT_DIR, "Innoflame_merged_sales_csv_source_with_L1_L2_L3_delivery_handling.xlsx"
)

MANUAL_COLS = [
    "source_excel_row",
    "update_key",
    "manual_update_status",
    "manual_product_group_l1_code",
    "manual_product_group_l1_name",
    "manual_product_group_l2_code",
    "manual_product_group_l2_name",
    "manual_product_group_l3_code",
    "manual_product_group_l3_name",
    "manual_update_note",
]

NUMERIC_COLS = {"id", "sales", "amount", "order", "accountid", "totalprice"}
GROUP_COLS = [
    "product_group_l1_code",
    "product_group_l1_name",
    "product_group_l2_code",
    "product_group_l2_name",
    "product_group_l3_code",
    "product_group_l3_name",
]


def parse_cell(header, value):
    if value == "":
        return None
    if header in NUMERIC_COLS:
        try:
            number = float(value)
            if number.is_integer():
                return int(number)
            return number
        except ValueError:
            return value
    return value


def is_blank(value):
    return value is None or str(value).strip() == ""


def has_group(row):
    return not is_blank(row.get("product_group_l3_code")) and not is_blank(
        row.get("product_group_l3_name")
    )


def style_sheet(ws, widths):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_xlsx = XLSX_PATH.replace(
        ".xlsx", f".backup_before_csv_sync_{timestamp}.xlsx"
    )
    temp_xlsx = XLSX_PATH.replace(".xlsx", f".tmp_sync_{timestamp}.xlsx")
    audit_json = os.path.join(OUTPUT_DIR, "xlsx_sync_after_group_fill_audit.json")

    shutil.copy2(XLSX_PATH, backup_xlsx)

    wb = Workbook(write_only=False)
    ws_in = wb.active
    ws_in.title = "in"
    ws_missing = wb.create_sheet("SKU_puuttuu_tyotaulu")
    ws_summary = wb.create_sheet("SKU_puuttuu_summary")

    total_rows = 0
    missing_sku_rows = 0
    missing_sku_sales = 0.0
    missing_group_rows = 0
    with_group_rows = 0
    method_counts = {}
    missing_names = {}

    with open(CSV_PATH, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        headers = reader.fieldnames or []
        ws_in.append(headers)
        ws_missing.append(MANUAL_COLS + headers)

        for data_row_number, row in enumerate(reader, start=2):
            total_rows += 1
            parsed = [parse_cell(header, row.get(header, "")) for header in headers]
            ws_in.append(parsed)

            method = row.get("product_group_match_method", "") or ""
            method_counts[method] = method_counts.get(method, 0) + 1

            sales_value = parse_cell("sales", row.get("sales", ""))
            sales_number = sales_value if isinstance(sales_value, (int, float)) else 0.0

            if has_group(row):
                with_group_rows += 1
            else:
                missing_group_rows += 1

            if is_blank(row.get("productcode")):
                missing_sku_rows += 1
                missing_sku_sales += sales_number
                name = row.get("name", "") or ""
                missing_names[name] = missing_names.get(name, 0) + 1
                update_key = f"{row.get('source_file', '')}|{row.get('id', '')}|{data_row_number}"
                ws_missing.append(
                    [
                        data_row_number,
                        update_key,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ]
                    + parsed
                )

    repeated_missing_names = sum(1 for count in missing_names.values() if count > 1)
    rows_in_repeated_missing_names = sum(
        count for count in missing_names.values() if count > 1
    )

    summary_rows = [
        ("source_csv", CSV_PATH),
        ("target_xlsx", XLSX_PATH),
        ("backup_xlsx", backup_xlsx),
        ("total_rows", total_rows),
        ("rows_with_l3_group", with_group_rows),
        ("rows_missing_l3_group", missing_group_rows),
        ("missing_sku_rows", missing_sku_rows),
        ("missing_sku_sales_eur", round(missing_sku_sales, 2)),
        ("unique_missing_sku_names", len(missing_names)),
        ("repeated_missing_sku_names", repeated_missing_names),
        ("rows_in_repeated_missing_sku_names", rows_in_repeated_missing_names),
    ]
    ws_summary.append(["metric", "value"])
    for row in summary_rows:
        ws_summary.append(row)
    ws_summary.append([])
    ws_summary.append(["match_method", "rows"])
    for method, count in sorted(method_counts.items(), key=lambda item: (-item[1], item[0])):
        ws_summary.append([method or "(blank)", count])

    style_sheet(
        ws_in,
        {
            1: 20,
            4: 18,
            5: 18,
            7: 24,
            9: 24,
            11: 24,
            12: 30,
            14: 34,
            18: 18,
        },
    )
    style_sheet(
        ws_missing,
        {
            1: 16,
            2: 34,
            3: 22,
            5: 26,
            7: 26,
            9: 26,
            10: 26,
            20: 24,
            24: 34,
        },
    )
    style_sheet(ws_summary, {1: 34, 2: 90})

    wb.save(temp_xlsx)

    # Verify the created workbook is readable before replacing the original.
    check = load_workbook(temp_xlsx, read_only=True, data_only=True)
    sheet_info = {ws.title: {"rows": ws.max_row, "cols": ws.max_column} for ws in check.worksheets}
    check.close()

    os.replace(temp_xlsx, XLSX_PATH)

    audit = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "source_csv": CSV_PATH,
        "target_xlsx": XLSX_PATH,
        "backup_xlsx": backup_xlsx,
        "rows_total": total_rows,
        "rows_with_l3_group": with_group_rows,
        "rows_missing_l3_group": missing_group_rows,
        "missing_sku_rows": missing_sku_rows,
        "missing_sku_sales_eur": round(missing_sku_sales, 2),
        "sheet_info": sheet_info,
        "method_counts": method_counts,
    }
    with open(audit_json, "w", encoding="utf-8") as fh:
        json.dump(audit, fh, ensure_ascii=False, indent=2)
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
